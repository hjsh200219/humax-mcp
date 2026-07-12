"""openpyxl wrappers with schema/lock validation."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Literal

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from ..schemas import bp26, raw_bp26
from . import errors

_SCHEMA_MODULES: dict[str, Any] = {"bp26": bp26, "raw_bp26": raw_bp26}

_AUTO_DETECT_MIN_MATCHES = 5
_AUTO_DETECT_SCAN_ROWS = 10


def assert_xlsx_path(path: str | Path) -> Path:
    p = Path(path)
    if not p.exists():
        raise errors.FileNotFound(f"파일을 찾을 수 없습니다: {p}")
    if p.suffix.lower() != ".xlsx":
        raise errors.FileNotFound(f"xlsx 확장자가 아닙니다: {p}")
    return p


def load_workbook_safe(
    path: str | Path,
    *,
    data_only: bool = True,
    read_only: bool = False,
) -> Workbook:
    p = assert_xlsx_path(path)
    try:
        return load_workbook(p, data_only=data_only, read_only=read_only)
    except PermissionError as exc:
        raise errors.FileLocked(
            f"파일이 다른 프로그램에서 열려 있습니다. Excel을 닫고 다시 시도하세요. ({p})"
        ) from exc
    except InvalidFileException as exc:
        raise errors.FileNotFound(f"유효한 xlsx 파일이 아닙니다: {p}") from exc


def get_sheet(wb: Workbook, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        raise errors.SheetNotFound(
            f"시트를 찾을 수 없습니다: {sheet_name}. 사용 가능: {wb.sheetnames}"
        )
    return wb[sheet_name]


def _count_schema_matches(row: tuple, schema_module: Any) -> int:
    """Count how many cells in a row match keys in the schema's COLUMN_MAP."""
    keys = set(schema_module.COLUMN_MAP.keys())
    return sum(1 for cell in row if cell is not None and str(cell) in keys)


def _dedup_headers(headers: list[str]) -> list[str]:
    """Deduplicate headers by appending __N suffixes."""
    seen: dict[str, int] = {}
    deduped: list[str] = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            deduped.append(f"{h}__{seen[h]}")
        else:
            seen[h] = 0
            deduped.append(h)
    return deduped


def worksheet_to_dataframe(
    ws,
    *,
    header_row: int | None = None,
    schema_module: str = "bp26",
    normalize: bool = True,
) -> pd.DataFrame:
    """Read a worksheet into a DataFrame.

    Behavior matrix:
    - Default (header_row=None, schema_module="bp26"):
        Backward-compat path. Uses row 1 as header directly. No auto-detect,
        no SchemaMismatch raise (preserves v0.1/v0.1.1 callers behavior).
    - schema_module="raw_bp26", header_row=None:
        Auto-detect: scan rows 1.._AUTO_DETECT_SCAN_ROWS for the row with the
        most schema-key matches. Pick the first row with >= _AUTO_DETECT_MIN_MATCHES.
        Raise SchemaMismatch if none qualify.
    - header_row=N (explicit):
        Use row N as header. Skip auto-detect.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()

    mod = _SCHEMA_MODULES.get(schema_module, bp26)

    # Determine which row is the header (0-indexed internally)
    if header_row is not None:
        # Explicit: use the given 1-indexed row number
        hdr_idx = header_row - 1
    elif schema_module == "bp26":
        # Backward-compat: always row 1
        hdr_idx = 0
    else:
        # Auto-detect: scan up to _AUTO_DETECT_SCAN_ROWS rows
        hdr_idx = None
        scan_limit = min(_AUTO_DETECT_SCAN_ROWS, len(rows))
        for i in range(scan_limit):
            if _count_schema_matches(rows[i], mod) >= _AUTO_DETECT_MIN_MATCHES:
                hdr_idx = i
                break
        if hdr_idx is None:
            raise errors.SchemaMismatch(
                f"헤더 행 자동감지 실패: 행 1-{scan_limit} 중 schema '{schema_module}' "
                f"키 {_AUTO_DETECT_MIN_MATCHES}개 이상 일치하는 행 없음. "
                "header_row= 로 직접 지정하거나 올바른 파일인지 확인하세요."
            )

    raw_headers = [str(h) if h is not None else "" for h in rows[hdr_idx]]

    if normalize:
        headers = mod.normalize_headers(raw_headers)
    else:
        headers = raw_headers

    deduped = _dedup_headers(headers)
    data = rows[hdr_idx + 1 :]
    return pd.DataFrame(data, columns=deduped)


def validate_schema(
    headers: list[str],
    *,
    require_allocation_cols: bool = False,
    strict: bool = False,
) -> list[str]:
    diffs = bp26.validate_headers(headers)
    missing = [d for d in diffs if d.startswith("MISSING:")]
    if strict and missing:
        raise errors.SchemaMismatch(
            f"파일 헤더가 스키마 v{bp26.SCHEMA_VERSION}과 일치하지 않습니다. "
            f"변경된 컬럼: {missing}. schemas/bp26.py를 업데이트하세요."
        )
    if require_allocation_cols:
        needed = set(bp26.ALLOCATION_RATE_COLUMNS.keys())
        present = set(headers)
        absent = needed - present
        if absent:
            raise errors.SchemaMismatch(
                f"C30-C34 배부율 컬럼이 누락되었습니다. 스키마 v{bp26.SCHEMA_VERSION} 확인 필요. "
                f"누락: {sorted(absent)}"
            )
    return diffs


def detect_source_format(ws) -> tuple[Literal["aggregated", "raw"], int]:
    """Detect whether a worksheet is aggregated (bp26) or raw (raw_bp26).

    Returns (format_name, 1-indexed header row number).

    Raises SchemaMismatch if no row in 1-10 matches any known schema.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise errors.SchemaMismatch(
            "Source format undetectable: no row in 1-10 matches any known schema"
        )

    # Try aggregated: bp26 matches on row 1
    if _count_schema_matches(rows[0], bp26) >= _AUTO_DETECT_MIN_MATCHES:
        return ("aggregated", 1)

    # Try raw: scan rows 1-10 for raw_bp26 matches
    scan_limit = min(_AUTO_DETECT_SCAN_ROWS, len(rows))
    for i in range(scan_limit):
        if _count_schema_matches(rows[i], raw_bp26) >= _AUTO_DETECT_MIN_MATCHES:
            return ("raw", i + 1)

    raise errors.SchemaMismatch(
        "Source format undetectable: no row in 1-10 matches any known schema"
    )


def cell_to_value(value: Any) -> Any:
    """Normalize openpyxl cell value (passthrough)."""
    return value
