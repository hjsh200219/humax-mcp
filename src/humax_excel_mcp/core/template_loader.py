"""Load golden template xlsx + sidecar metadata. Plan §2."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from . import errors
from .template_bindings import TemplateBinding, get_binding


def load_template(
    template_path: str | Path,
    template_type: str,
) -> tuple[Workbook, dict, TemplateBinding]:
    """Load template workbook + sidecar metadata + binding. Validates structure."""
    tpath = Path(template_path)
    if not tpath.exists():
        raise errors.TemplateNotFound(
            f"템플릿 파일을 찾을 수 없습니다: {tpath}"
        )
    if tpath.suffix.lower() != ".xlsx":
        raise errors.TemplateMalformed(
            f"템플릿은 xlsx여야 합니다: {tpath}"
        )

    sidecar_path = tpath.with_suffix(".template.json")
    if not sidecar_path.exists():
        raise errors.TemplateMalformed(
            f"sidecar JSON이 없습니다: {sidecar_path}"
        )
    try:
        metadata = json.loads(sidecar_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise errors.TemplateMalformed(
            f"sidecar JSON 파싱 실패: {sidecar_path} — {exc}"
        ) from exc

    binding = get_binding(template_type)

    if metadata.get("template_type") != template_type:
        raise errors.TemplateMalformed(
            f"sidecar template_type 불일치: 파일={metadata.get('template_type')!r}, "
            f"요청={template_type!r}"
        )
    if metadata.get("schema_version") != binding.schema_version:
        raise errors.TemplateMalformed(
            f"schema_version 불일치: 파일={metadata.get('schema_version')!r}, "
            f"binding={binding.schema_version!r}"
        )

    try:
        wb = load_workbook(tpath, keep_vba=False, keep_links=True)
    except PermissionError as exc:
        raise errors.FileLocked(
            f"파일이 다른 프로그램에서 열려 있습니다: {tpath}"
        ) from exc

    expected_sheets = set(metadata.get("sheet_names", []))
    actual_sheets = set(wb.sheetnames)
    if expected_sheets and expected_sheets != actual_sheets:
        missing = expected_sheets - actual_sheets
        extra = actual_sheets - expected_sheets
        raise errors.TemplateMalformed(
            f"sheet 구조 불일치 — missing={sorted(missing)} extra={sorted(extra)}"
        )

    for sb in binding.sheets:
        if sb.sheet_name not in actual_sheets:
            raise errors.TemplateMalformed(
                f"binding 시트 '{sb.sheet_name}' 가 템플릿에 없습니다."
            )

    return wb, metadata, binding
