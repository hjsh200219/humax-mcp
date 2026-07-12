"""verify_sums tool — PRD §4.2 + accuracy-speed PRD US-A1/A3/S-2.

계층 검증 규칙:
- 총합계: 총합계 행 합 vs 소조직 상세 합.
- 사업부/대조직/중조직: 각 소계 행의 비어있지 않은 org 컬럼 prefix로
  소조직 상세를 스코핑해 합산 대조. 소계 행이 없으면 SKIPPED.
- 소조직: 행 내부 일관성 — 연간 실적 vs 월별 실적 합.
"""

from __future__ import annotations

from pathlib import Path
from typing import Literal

import pandas as pd
from openpyxl.utils import get_column_letter

from ..core import artifact_hints as ah
from ..core import errors, excel_io, workbook_cache
from ..schemas.responses import (
    Anomaly,
    FormulaWarning,
    LevelResult,
    VerifyResult,
    VerifySummary,
)

DEFAULT_LEVELS = ["총합계", "사업부", "대조직", "중조직", "소조직"]
ANOMALY_THRESHOLD_MILLION = 10.0

_ORG_COLS = ["org_l1", "org_l2", "org_l3"]
_MONTHLY_ACTUAL_COLS = [f"m{m:02d}_actual" for m in range(1, 13)]
_SUBTOTAL_DIVISIONS = ("총합계", "사업부", "대조직", "중조직")


def _skipped(level: str, detail: str) -> LevelResult:
    return LevelResult(
        level=level,
        expected=0.0,
        actual=0.0,
        difference=0.0,
        status="SKIPPED",
        detail=detail,
    )


def _norm_org(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.strip()


def _check_total(
    df: pd.DataFrame, detail: pd.DataFrame, actual_col: str, tolerance: float
) -> LevelResult:
    rows = df[df["division"] == "총합계"]
    if rows.empty:
        return _skipped("총합계", "총합계 행 없음")
    expected = float(detail[actual_col].sum())
    actual = float(rows[actual_col].sum())
    diff = actual - expected
    status = "PASS" if abs(diff) <= tolerance else "FAIL"
    return LevelResult(
        level="총합계",
        expected=expected,
        actual=actual,
        difference=diff,
        status=status,
        detail=(None if status == "PASS" else f"총합계 차이 {diff}"),
    )


def _check_rollup(
    level: str, df: pd.DataFrame, detail: pd.DataFrame, actual_col: str, tolerance: float
) -> LevelResult:
    rows = df[df["division"] == level]
    if rows.empty:
        return _skipped(level, "해당 계층 소계 행 없음")
    org_norm = {c: (_norm_org(detail[c]) if c in detail.columns else None) for c in _ORG_COLS}
    exp_sum = act_sum = 0.0
    breaches = 0
    scoped_rows = 0
    for _, row in rows.iterrows():
        mask = pd.Series(True, index=detail.index)
        scoped = False
        for c in _ORG_COLS:
            raw_v = row.get(c)
            v = "" if raw_v is None or pd.isna(raw_v) else str(raw_v).strip()
            if v and org_norm[c] is not None:
                mask &= org_norm[c] == v
                scoped = True
        if not scoped:
            continue
        scoped_rows += 1
        expected = float(detail.loc[mask, actual_col].sum())
        actual = float(row[actual_col])
        exp_sum += expected
        act_sum += actual
        if abs(actual - expected) > tolerance:
            breaches += 1
    if scoped_rows == 0:
        return _skipped(level, "소계 행에 조직 키 없음 — 스코핑 불가")
    status = "FAIL" if breaches else "PASS"
    return LevelResult(
        level=level,
        expected=exp_sum,
        actual=act_sum,
        difference=act_sum - exp_sum,
        status=status,
        detail=(None if not breaches else f"{breaches}개 소계 행이 하위 소조직 합과 불일치"),
    )


def _check_detail_consistency(
    detail: pd.DataFrame, actual_col: str, tolerance: float
) -> LevelResult:
    monthly = [c for c in _MONTHLY_ACTUAL_COLS if c in detail.columns]
    if not monthly:
        return _skipped("소조직", "월별 실적 컬럼 없음")
    monthly_sum = detail[monthly].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1)
    annual = detail[actual_col]
    breaches = int(((annual - monthly_sum).abs() > tolerance).sum())
    expected = float(monthly_sum.sum())
    actual = float(annual.sum())
    status = "FAIL" if breaches else "PASS"
    return LevelResult(
        level="소조직",
        expected=expected,
        actual=actual,
        difference=actual - expected,
        status=status,
        detail=(None if not breaches else f"{breaches}개 행 월별 실적 합≠연간 실적"),
    )


def _collect_anomalies(detail: pd.DataFrame, budget_col: str, actual_col: str) -> list[Anomaly]:
    anomalies: list[Anomaly] = []
    for idx, row in detail.iterrows():
        budget = float(row[budget_col])
        actual = float(row[actual_col])
        variance_million = (actual - budget) / 1_000_000.0
        if abs(variance_million) < ANOMALY_THRESHOLD_MILLION:
            continue
        account = str(row.get("gl_account_name", ""))
        org = " > ".join(str(row.get(c, "") or "") for c in ("org_l1", "org_l2", "org_l3"))
        sign = "+" if variance_million >= 0 else "-"
        code = str(row.get("gl_account", ""))[:2]
        anomalies.append(
            Anomaly(
                row_index=int(idx) + 2,
                org=org,
                account=account,
                budget=budget,
                actual=actual,
                variance=actual - budget,
                flag="LARGE_VARIANCE",
                suggested_comment=f"{code} {account} {sign}{abs(int(variance_million))}백만",
            )
        )
    return anomalies


def _scan_formula_warnings(
    file_path: str, sheet_name: str
) -> tuple[list[FormulaWarning], str | None]:
    """Read-only 수식 스캔. 실패는 침묵하지 않고 (warnings, error_msg)로 반환."""
    try:
        wb = excel_io.load_workbook_safe(file_path, data_only=False, read_only=True)
        ws = excel_io.get_sheet(wb, sheet_name)
        warnings: list[FormulaWarning] = []
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if headers is None:
            wb.close()
            return [], None
        subtotal_cols = {
            i
            for i, h in enumerate(headers)
            if h in ("연간 예산", "연간 실적") or (isinstance(h, str) and "누계" in h)
        }
        for r_idx, row in enumerate(rows_iter, start=2):
            if not row or row[0] not in _SUBTOTAL_DIVISIONS:
                continue
            for c_idx in subtotal_cols:
                if c_idx >= len(row):
                    continue
                value = row[c_idx]
                if value is None or value == "":
                    continue
                if isinstance(value, str) and value.startswith("="):
                    continue
                warnings.append(
                    FormulaWarning(
                        cell=f"{get_column_letter(c_idx + 1)}{r_idx}",
                        expected_formula="=SUM(...)",
                        current_state="hard_coded_value",
                        warning="수식이 값으로 덮어쓰여져 있습니다.",
                    )
                )
        wb.close()
        return warnings, None
    except Exception as exc:
        return [], f"수식 검사 실패: {type(exc).__name__}: {exc}"


async def verify_sums(
    file_path: str,
    sheet_name: str,
    *,
    levels: list[str] | None = None,
    tolerance: float = 0.01,
    check_formulas: bool = True,
    render_format: Literal["excel", "live_artifact", "both"] = "excel",
) -> VerifyResult:
    """Verify hierarchical sums per PRD §4.2 (실계층 rollup)."""
    if levels is None:
        levels = DEFAULT_LEVELS

    df = workbook_cache.get_dataframe(file_path, sheet_name)
    if df.empty:
        raise errors.ParseError("조직 계층 구조를 인식할 수 없습니다. 시트 형식을 확인하세요.")
    if "division" not in df.columns:
        raise errors.ParseError("'구분' 컬럼을 찾을 수 없습니다.")

    budget_col = "annual_budget" if "annual_budget" in df.columns else None
    actual_col = "annual_actual" if "annual_actual" in df.columns else None
    if budget_col is None or actual_col is None:
        raise errors.SubtotalNotFound("연간 예산/실적 컬럼이 없어 합계 검증 불가합니다.")

    df[budget_col] = pd.to_numeric(df[budget_col], errors="coerce").fillna(0)
    df[actual_col] = pd.to_numeric(df[actual_col], errors="coerce").fillna(0)

    detail = df[df["division"] == "소조직"].copy()
    if detail.empty:
        raise errors.SubtotalNotFound("소조직 상세 행을 찾을 수 없습니다.")

    level_results: list[LevelResult] = []
    for level in levels:
        if level == "총합계":
            level_results.append(_check_total(df, detail, actual_col, tolerance))
        elif level == "소조직":
            level_results.append(_check_detail_consistency(detail, actual_col, tolerance))
        else:
            level_results.append(_check_rollup(level, df, detail, actual_col, tolerance))

    anomalies = _collect_anomalies(detail, budget_col, actual_col)

    formula_warnings: list[FormulaWarning] = []
    formula_error: str | None = None
    if check_formulas:
        formula_warnings, formula_error = _scan_formula_warnings(file_path, sheet_name)

    passed = sum(1 for r in level_results if r.status == "PASS")
    failed = sum(1 for r in level_results if r.status == "FAIL")
    skipped = sum(1 for r in level_results if r.status == "SKIPPED")
    summary = VerifySummary(
        total_checks=len(level_results),
        passed=passed,
        failed=failed,
        warnings=len(formula_warnings) + (1 if formula_error else 0),
        skipped=skipped,
    )

    hints = ah.maybe_hints(
        render_format,
        artifact_type="verification_result",
        title=f"{sheet_name} 합계 검증 결과",
        preferred_chart="tree",
    )

    return VerifyResult(
        summary=summary,
        level_results=level_results,
        anomalies=anomalies[:50],
        formula_warnings=formula_warnings,
        metadata={
            "file_path": str(Path(file_path)),
            "sheet_name": sheet_name,
            "levels_checked": levels,
            "tolerance": tolerance,
            "anomaly_threshold_million": ANOMALY_THRESHOLD_MILLION,
            "anomaly_total_count": len(anomalies),
            "formula_check_error": formula_error,
        },
        render_format=render_format,
        artifact_hints=hints,
    )
