"""apply_golden_template — PRD §4.9 / plan §5 US-016. Deterministic golden-template engine."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Literal

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from ..core import aggregator, errors, excel_io
from ..core import artifact_hints as ah
from ..core import backup as backup_mod
from ..core.excel_io import detect_source_format
from ..core.template_bindings import SheetBinding
from ..core.template_loader import load_template
from ..schemas.responses import (
    ApplyTemplateResult,
    TemplateBindingSummary,
    WriteVerification,
)


def _filter_and_sort(df: pd.DataFrame, binding: SheetBinding) -> pd.DataFrame:
    sel = binding.row_selection
    work = df
    if sel.filter_column and sel.filter_column in work.columns:
        work = work[work[sel.filter_column].isin(sel.filter_values)]
    if sel.sort_by:
        cols = [c for c in sel.sort_by if c in work.columns]
        if cols:
            work = work.sort_values(by=cols)
    return work.reset_index(drop=True)


def _populate_sheet(ws, df: pd.DataFrame, binding: SheetBinding) -> TemplateBindingSummary:
    cells_to_populate = 0
    formulas_preserved = 0
    rows_matched = 0
    rows_unmatched = 0

    work = _filter_and_sort(df, binding)

    iter_rows = iter(work.to_dict(orient="records"))
    skip_set = set(binding.skip_formula_rows)

    for row_idx in range(binding.data_start_row, binding.data_end_row + 1):
        if row_idx in skip_set:
            continue
        try:
            data_row = next(iter_rows)
            rows_matched += 1
        except StopIteration:
            rows_unmatched += 1
            continue

        for col_letter, source_key in binding.column_map.items():
            cell = ws[f"{col_letter}{row_idx}"]
            if isinstance(cell, MergedCell):
                continue
            existing = cell.value
            if isinstance(existing, str) and existing.startswith("="):
                formulas_preserved += 1
                continue
            value = data_row.get(source_key)
            if value is None or (isinstance(value, float) and pd.isna(value)):
                cell.value = None
            else:
                cell.value = value
            cells_to_populate += 1

    return TemplateBindingSummary(
        sheet_name=binding.sheet_name,
        cells_to_populate=cells_to_populate,
        formulas_preserved=formulas_preserved,
        rows_matched=rows_matched,
        rows_unmatched=rows_unmatched,
    )


async def apply_golden_template(
    source_file: str,
    template_path: str,
    template_type: Literal["humax_allocation", "humax_account", "evcs_account"],
    output_path: str,
    *,
    month: int,
    dry_run: bool = False,
    render_format: Literal["excel", "live_artifact", "both"] = "excel",
    source_format: Literal["auto", "raw", "aggregated"] = "auto",
    expand_evcs: bool = False,
) -> ApplyTemplateResult:
    """Apply a golden template to source 26BP data — deterministic, zero design drift."""
    if not 1 <= month <= 12:
        raise errors.InvalidMonth(f"month는 1-12 범위여야 합니다. 입력: {month}")

    src = excel_io.assert_xlsx_path(source_file)
    out = Path(output_path)
    if out.resolve() == src.resolve():
        raise errors.OverwriteOriginalForbidden(
            "output_path가 원본 source_file과 동일합니다. 원본 직접 덮어쓰기는 허용되지 않습니다."
        )
    if not out.parent.exists():
        raise errors.WritePermissionDenied(f"출력 디렉터리가 없습니다: {out.parent}")

    template_wb, template_meta, binding = load_template(template_path, template_type)

    backup_path = None
    if not dry_run:
        backup_path = backup_mod.create_backup(src)

    # read_only=True: source is input-only, never mutated. Avoids ~2GB / multi-minute
    # full-formula evaluation pass on large real-data files (15k+ rows).
    src_wb = excel_io.load_workbook_safe(src, data_only=True, read_only=True)
    sheet_name = "예산+실적"
    if sheet_name not in src_wb.sheetnames:
        raise errors.SheetNotFound(
            f"source_file에 '{sheet_name}' 시트가 없습니다. 사용 가능: {src_wb.sheetnames}"
        )
    src_ws = src_wb[sheet_name]

    # Determine source format and load DataFrame accordingly
    if source_format == "auto":
        fmt, hdr_row = detect_source_format(src_ws)
    elif source_format == "raw":
        fmt, hdr_row = "raw", None
    elif source_format == "aggregated":
        fmt, hdr_row = "aggregated", 1
    else:
        raise errors.SchemaMismatch(f"Unknown source_format: {source_format}")

    if fmt == "aggregated":
        df = excel_io.worksheet_to_dataframe(src_ws)
        # Validate that the df actually has aggregated-format columns
        if source_format == "aggregated":
            required_keys = {"company", "gl_account", "cost_center"}
            present = set(df.columns)
            if len(required_keys & present) < 3:
                raise errors.SchemaMismatch(
                    f"source_format='aggregated' が指定されましたが、bp26スキーマのキー ({sorted(required_keys)}) が "
                    f"列に見つかりません。実際の列: {list(df.columns)[:10]}"
                )
    elif fmt == "raw":
        raw_df = excel_io.worksheet_to_dataframe(
            src_ws,
            header_row=hdr_row,
            schema_module="raw_bp26",
        )
        agg_result = aggregator.aggregate_to_bp26(
            raw_df,
            target_month=month,
            expand_evcs=expand_evcs,
        )
        df = agg_result.df
    else:
        raise errors.SchemaMismatch(f"Unknown source_format: {fmt}")

    sheets_processed: list[TemplateBindingSummary] = []
    for sb in binding.sheets:
        ws = template_wb[sb.sheet_name]
        summary = _populate_sheet(ws, df, sb)
        sheets_processed.append(summary)

    verification = WriteVerification(verified=True)

    if not dry_run:
        template_wb.save(out)
        check_wb = load_workbook(out, data_only=False)
        mismatches: list[dict] = []
        # Spot-check the first populated row of each bound sheet against expected source data
        for sb in binding.sheets:
            ws_check = check_wb[sb.sheet_name]
            work = _filter_and_sort(df, sb)
            if work.empty:
                continue
            first_record = work.to_dict(orient="records")[0]
            for col_letter, source_key in sb.column_map.items():
                cell = ws_check[f"{col_letter}{sb.data_start_row}"]
                if isinstance(cell, MergedCell):
                    continue
                existing = cell.value
                if isinstance(existing, str) and existing.startswith("="):
                    continue
                expected = first_record.get(source_key)
                if expected is None or (isinstance(expected, float) and pd.isna(expected)):
                    if cell.value is not None:
                        mismatches.append({
                            "sheet": sb.sheet_name,
                            "cell": f"{col_letter}{sb.data_start_row}",
                            "expected": None,
                            "got": cell.value,
                        })
                else:
                    if cell.value != expected:
                        mismatches.append({
                            "sheet": sb.sheet_name,
                            "cell": f"{col_letter}{sb.data_start_row}",
                            "expected": expected,
                            "got": cell.value,
                        })
        if mismatches:
            verification = WriteVerification(verified=False, mismatches=mismatches)
            raise errors.VerificationFailed(
                f"쓰기 후 검증 실패: {mismatches[:3]}. 백업에서 복구하세요."
            )

    hints = ah.maybe_hints(
        render_format,
        artifact_type="dashboard",
        title=f"{template_type} 산출물 (M{month:02d})",
        preferred_chart="bar",
    )

    metadata: dict[str, Any] = {
        "source_file": str(src),
        "template_path": str(template_path),
        "template_type": template_type,
        "template_schema_version": template_meta.get("schema_version"),
        "month": month,
        "sheets_in_binding": len(binding.sheets),
        "sheets_in_template": len(template_wb.sheetnames),
    }

    return ApplyTemplateResult(
        dry_run=dry_run,
        template_type=template_type,
        output_path=str(out) if not dry_run else None,
        backup_path=str(backup_path) if backup_path else None,
        sheets_processed=sheets_processed,
        verification=verification,
        metadata=metadata,
        render_format=render_format,
        artifact_hints=hints,
    )
