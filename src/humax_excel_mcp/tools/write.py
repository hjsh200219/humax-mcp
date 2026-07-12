"""write_cells tool — PRD §4.3. Critical safety: backup-always, no overwrite, dry-run, post-verify."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

from ..core import artifact_hints as ah
from ..core import backup as backup_mod
from ..core import errors, excel_io, workbook_cache
from ..schemas.requests import CellUpdate
from ..schemas.responses import (
    WriteApplied,
    WriteResult,
    WriteSkipped,
    WriteSummary,
    WriteVerification,
    WriteWarning,
)

CELL_RE = re.compile(r"^[A-Z]{1,3}[1-9][0-9]*$")
MAX_UPDATES = 5000


def _default_output_path(file_path: str | Path) -> Path:
    p = Path(file_path)
    return p.with_name(f"{p.stem}_edited{p.suffix}")


async def write_cells(
    file_path: str,
    sheet_name: str,
    updates: list[CellUpdate] | list[dict],
    *,
    output_path: str | None = None,
    dry_run: bool = False,
    render_format: Literal["excel", "live_artifact", "both"] = "excel",
) -> WriteResult:
    """Edit specified cells with backup + dry-run + post-verify."""
    parsed: list[CellUpdate] = []
    for u in updates:
        parsed.append(u if isinstance(u, CellUpdate) else CellUpdate(**u))

    if len(parsed) == 0:
        raise errors.TooManyUpdates("EMPTY_UPDATES: 업데이트가 비어 있습니다.")
    if len(parsed) > MAX_UPDATES:
        raise errors.TooManyUpdates(
            f"최대 {MAX_UPDATES}개 셀까지 편집 가능합니다. 요청: {len(parsed)}"
        )

    src = excel_io.assert_xlsx_path(file_path)

    out = Path(output_path) if output_path else _default_output_path(src)
    if out.resolve() == src.resolve():
        raise errors.OverwriteOriginalForbidden(
            "output_path가 원본 file_path와 동일합니다. 원본 직접 덮어쓰기는 허용되지 않습니다."
        )

    if not out.parent.exists():
        raise errors.WritePermissionDenied(f"출력 디렉터리가 없습니다: {out.parent}")

    backup_path = None
    if not dry_run:
        backup_path = backup_mod.create_backup(src)

    wb = excel_io.load_workbook_safe(src, data_only=False)
    ws = excel_io.get_sheet(wb, sheet_name)

    applied: list[WriteApplied] = []
    skipped: list[WriteSkipped] = []
    warnings_list: list[WriteWarning] = []

    # ── AB=None 사전 감지: 지급수수료 시트에서 예산 수식 누락 행 경고 ──────────
    # AB열(col 28)=4월예산은 SUMIFS 수식이어야 함.
    # B열(CC)과 E열(GL)이 모두 채워진 행에서 AB가 None이면 합계 불일치 원인이 됨.
    if sheet_name == "지급수수료":
        AB_COL, B_COL, E_COL = 28, 2, 5
        missing_ab: list[int] = []
        for row in ws.iter_rows(min_row=2):
            b_val = row[B_COL - 1].value
            e_val = row[E_COL - 1].value
            ab_val = row[AB_COL - 1].value
            if b_val and e_val and (ab_val is None or ab_val == ""):
                missing_ab.append(row[0].row)
        if missing_ab:
            rows_str = ", ".join(str(r) for r in missing_ab[:20])
            warnings_list.append(
                WriteWarning(
                    cell="AB열",
                    message=(
                        f"⚠️ [예산수식 누락 감지] CC+GL 값이 있으나 AB(4월예산)가 비어 있는 행: "
                        f"{rows_str}행. "
                        "write_cells 실행 전 해당 행에 SUMIFS 수식을 먼저 추가하세요. "
                        "누락 시 예산 합계에 차이가 발생합니다."
                    ),
                )
            )
    # ─────────────────────────────────────────────────────────────────────────

    seen: dict[str, int] = {}
    for upd in parsed:
        seen[upd.cell] = seen.get(upd.cell, 0) + 1

    duplicates = {k for k, v in seen.items() if v > 1}
    for d in sorted(duplicates):
        warnings_list.append(WriteWarning(cell=d, message="중복 셀 — 마지막 값 적용"))

    for upd in parsed:
        cell = ws[upd.cell]
        old_value = cell.value
        is_formula = isinstance(old_value, str) and old_value.startswith("=")
        if is_formula and upd.skip_if_formula:
            skipped.append(
                WriteSkipped(
                    cell=upd.cell,
                    reason="formula_cell",
                    formula=str(old_value),
                )
            )
            continue
        if is_formula and not upd.skip_if_formula:
            warnings_list.append(
                WriteWarning(
                    cell=upd.cell,
                    message=f"수식 덮어쓰기: {old_value}",
                )
            )
        if not dry_run:
            cell.value = upd.value
        applied.append(WriteApplied(cell=upd.cell, old_value=old_value, new_value=upd.value))

    verification = WriteVerification(verified=True)

    if not dry_run:
        wb.save(out)
        workbook_cache.invalidate(out)
        wb_check = load_workbook(out, data_only=False)
        ws_check = wb_check[sheet_name]
        mismatches: list[dict] = []
        for a in applied:
            got = ws_check[a.cell].value
            if got != a.new_value:
                mismatches.append({"cell": a.cell, "expected": a.new_value, "got": got})
        if mismatches:
            verification = WriteVerification(verified=False, mismatches=mismatches)
            raise errors.VerificationFailed(
                f"쓰기 후 검증 실패: {mismatches[:3]}. 백업에서 복구하세요."
            )

    summary = WriteSummary(
        total_updates=len(parsed),
        applied=len(applied),
        skipped_formula=sum(1 for s in skipped if s.reason == "formula_cell"),
        skipped_invalid=sum(1 for s in skipped if s.reason != "formula_cell"),
        warnings=len(warnings_list),
    )

    if len(applied) > 100:
        applied = applied[:50] + applied[-50:]

    hints = ah.maybe_hints(
        render_format,
        artifact_type="diff_cards",
        title=f"{sheet_name} 셀 편집 {'(dry-run)' if dry_run else '결과'}",
    )

    return WriteResult(
        dry_run=dry_run,
        summary=summary,
        output_path=str(out) if not dry_run else None,
        backup_path=str(backup_path) if backup_path else None,
        applied=applied,
        skipped=skipped,
        warnings=warnings_list,
        verification=verification,
        render_format=render_format,
        artifact_hints=hints,
    )
