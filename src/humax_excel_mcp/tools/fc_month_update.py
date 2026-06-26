"""update_fc_month_report — US-023. FC 실적 보고서 당월 배부판 자동 생성.

raw data + 전월 FC 보고서를 입력받아 {월}(AC) 시트를 채우고, {월} 누계(AC)/{월} 누계(상세)
시트를 신규 생성한다. 표준 워크플로우는 docs/exec-plans/active/us-023-fc-month-update-tool.md 참고.
"""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Literal

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from ..core import artifact_hints as ah
from ..core import backup as backup_mod
from ..core import errors, excel_io
from ..core.fc_report_layout import (
    ALIAS_TO_ROW,
    CAT_MAP,
    CATS,
    COMMENT_ITEM_MIN,
    COMMENT_THRESHOLD,
    COMMENT_TOPN,
    LEAF_ROWS,
    RAW_COL_AMOUNT_KRW,
    RAW_COL_COMPANY,
    RAW_COL_DAEGYEJEONG_RE,
    RAW_COL_GUBUN,
    RAW_COL_MONTH,
    ROWS_CORP,
    ROWS_HQ,
    SEG_BLOCK_COLS,
    SEG_COLS_RAW,
    SEG_RAW_IDX,
)
from ..schemas.responses import (
    FcMonthAcSummary,
    FcMonthCumulativeSummary,
    FcMonthUpdateResult,
    FcMonthVerification,
)

SEGS = ["STB", "Mobility", "EVCS", "공통", "건물", "Shared"]
RAW_SHEET_NAME = "예산+실적"
_MONTH_RE = re.compile(r"^\d{1,2}월$")


def _parse_month(month: str) -> int:
    if not _MONTH_RE.match(month):
        raise errors.InvalidMonth(f"month는 'N월'(N=2~12) 형식이어야 합니다. 입력: {month!r}")
    n = int(month[:-1])
    if not 2 <= n <= 12:
        raise errors.InvalidMonth(
            f"month는 2~12 범위여야 합니다 (1월은 누계 체인의 시작점이라 본 도구 대상 아님). 입력: {month!r}"
        )
    return n


def _month_range(month: str) -> list[str]:
    n = int(month.replace("월", ""))
    return [f"{i}월" for i in range(1, n + 1)]


def _aggregate_org_actuals(
    raw_path: str, month: str
) -> tuple[dict[int, dict[str, float]], float, dict[str, float]]:
    """raw '예산+실적' 시트에서 month의 '실적' 행만 대조직x세그먼트로 집계."""
    wb = load_workbook(raw_path, data_only=True, read_only=True)
    if RAW_SHEET_NAME not in wb.sheetnames:
        raise errors.SheetNotFound(
            f"raw_path에 '{RAW_SHEET_NAME}' 시트가 없습니다. 사용 가능: {wb.sheetnames}"
        )
    ws = wb[RAW_SHEET_NAME]
    sums: dict[int, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    total_check = 0.0
    unmatched: dict[str, float] = defaultdict(float)
    matched_rows = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[RAW_COL_GUBUN] != "실적":
            continue
        if str(row[RAW_COL_MONTH]).strip() != month:
            continue
        matched_rows += 1
        org = str(row[10]).strip()
        amt = float(row[RAW_COL_AMOUNT_KRW] or 0)
        total_check += amt
        if org not in ALIAS_TO_ROW:
            unmatched[org] += amt
            continue
        r = ALIAS_TO_ROW[org]
        for col, idx in SEG_RAW_IDX.items():
            sums[r][col] += float(row[idx] or 0)
    wb.close()
    if matched_rows == 0:
        raise errors.EmptyResult(
            f"raw_path에 구분='실적', Month='{month}' 조건을 만족하는 행이 없습니다."
        )
    return sums, total_check, dict(unmatched)


def _restore_ac_formulas(ws_cur, ws_ref, month: str, prev_month: str) -> int:
    """{전월}(AC) 시트의 수식 문자열을 {당월}(AC)의 빈 셀에 복사 (월 문자열만 치환).

    템플릿이 매월 일관되게 누락시키는 BP블록 leaf row 수식 + 집계행 SUM 수식을 복원한다.
    """
    needle = f"'{prev_month}(BP)'"
    repl = f"'{month}(BP)'"
    restored = 0
    for r in range(1, 91):       # 3블록(BP/실적/Diff) 전체 = 행 1~90
        for c in range(1, 26):   # A~Y열
            rv = ws_ref.cell(r, c).value
            cv = ws_cur.cell(r, c).value
            if isinstance(rv, str) and rv.startswith("=") and cv is None:
                ws_cur.cell(r, c).value = rv.replace(needle, repl)
                restored += 1
    return restored


def _fill_ac_actuals(ws, sums: dict[int, dict[str, float]]) -> int:
    """실적 블록 leaf row(20개) x 세그먼트열(13개) = 260셀 기입."""
    written = 0
    for r in LEAF_ROWS:
        for col in SEG_RAW_IDX:
            v = sums.get(r, {}).get(col, 0.0)
            ws.cell(r, column_index_from_string(col)).value = round(v, 2)
            written += 1
    return written


def _load_raw_df(raw_path: str, months: list[str]) -> pd.DataFrame:
    wb = load_workbook(raw_path, data_only=True, read_only=True)
    if RAW_SHEET_NAME not in wb.sheetnames:
        raise errors.SheetNotFound(
            f"raw_path에 '{RAW_SHEET_NAME}' 시트가 없습니다. 사용 가능: {wb.sheetnames}"
        )
    ws = wb[RAW_SHEET_NAME]
    records = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[RAW_COL_GUBUN] not in ("예산", "실적"):
            continue
        if str(row[RAW_COL_MONTH]).strip() not in months:
            continue
        rec = {
            "gubun": row[RAW_COL_GUBUN],
            "month": row[RAW_COL_MONTH],
            "company": row[RAW_COL_COMPANY],
            "daegyejeong_re": (
                str(row[RAW_COL_DAEGYEJEONG_RE]).strip() if row[RAW_COL_DAEGYEJEONG_RE] else None
            ),
            "amount_krw": float(row[RAW_COL_AMOUNT_KRW] or 0),
        }
        for name, idx in SEG_COLS_RAW.items():
            rec[name] = row[idx] or 0
        records.append(rec)
    wb.close()
    df = pd.DataFrame(records)
    if df.empty:
        return df
    df["EVCS"] = df["EVCS_in"] + df["EVCS_out"]
    df["hb2"] = df["company"].apply(lambda c: "본사" if c == "HKR" else "법인")
    df["category"] = df["daegyejeong_re"].map(CAT_MAP).fillna("기타")
    return df


def _compute_totals(df: pd.DataFrame) -> dict[tuple[str, str, str], dict[str, float]]:
    results: dict[tuple[str, str, str], dict[str, float]] = {}
    for hb in ["본사", "법인"]:
        for gubun in ["예산", "실적"]:
            for cat in CATS:
                if df.empty:
                    sub = df
                else:
                    sub = df[(df.hb2 == hb) & (df.gubun == gubun) & (df.category == cat)]
                s = sub[SEGS].sum() if not sub.empty else pd.Series(0.0, index=SEGS)
                results[(cat, hb, gubun)] = {
                    "총합계": float(s.sum()),
                    **{k: float(s[k]) for k in SEGS},
                }
    return results


def _compute_account_diffs(
    df: pd.DataFrame,
) -> dict[tuple[str, str], dict[str, pd.Series]]:
    """카테고리 x 본사/법인 단위로 대계정(daegyejeong_re)별 (실적-예산) 차이를 세그먼트별로 계산.

    비고 코멘트에서 어떤 대계정이 차이를 주도했는지 식별하는 용도.
    """
    result: dict[tuple[str, str], dict[str, pd.Series]] = {}
    for hb in ["본사", "법인"]:
        for cat in CATS:
            sub = df[(df.hb2 == hb) & (df.category == cat)] if not df.empty else df
            seg_diffs: dict[str, pd.Series] = {}
            for seg in SEGS:
                if sub.empty:
                    seg_diffs[seg] = pd.Series(dtype=float)
                    continue
                actual = sub[sub.gubun == "실적"].groupby("daegyejeong_re")[seg].sum()
                budget = sub[sub.gubun == "예산"].groupby("daegyejeong_re")[seg].sum()
                seg_diffs[seg] = actual.sub(budget, fill_value=0.0)
            total = pd.Series(dtype=float)
            for s in seg_diffs.values():
                total = total.add(s, fill_value=0.0)
            seg_diffs["총합계"] = total
            result[(cat, hb)] = seg_diffs
    return result


def _format_comment(diff_series: pd.Series, force_top1: bool = False) -> str:
    """차이가 큰 대계정 상위 N개를 '{대계정} {부호}{금액}백만' 형태로 조합."""
    if diff_series is None or diff_series.empty:
        return ""
    ordered_idx = diff_series.abs().sort_values(ascending=False).index
    ordered = diff_series.reindex(ordered_idx)
    items = [(name, val) for name, val in ordered.items() if abs(val) >= COMMENT_ITEM_MIN]
    if not items and force_top1 and len(ordered) > 0:
        name, val = next(iter(ordered.items()))
        if val != 0:
            items = [(name, val)]
    items = items[:COMMENT_TOPN]
    parts = []
    for name, val in items:
        sign = "+" if val >= 0 else "△"
        amount_million = round(abs(val) / 1_000_000)
        parts.append(f"{name} {sign}{amount_million}백만")
    return " / ".join(parts)


def _build_cumulative_ac_sheet(wb, prev_month: str, month: str) -> tuple[object, int]:
    src = wb[f"{prev_month} 누계(AC)"]
    new = wb.copy_worksheet(src)
    new.title = f"{month} 누계(AC)"
    new.sheet_state = "visible"
    needle = f"'{prev_month}(AC)'!"
    n_modified = 0
    for row in new.iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=") and needle in v):
                continue
            tail = f"{needle}{cell.coordinate}"
            if not v.endswith(tail):
                continue  # 예상 패턴과 다른 셀은 건드리지 않음 (안전장치)
            cell.value = v + f"+'{month}(AC)'!{cell.coordinate}"
            n_modified += 1
    for row in new.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and not v.startswith("=") and prev_month in v:
                cell.value = v.replace(prev_month, month)
    src.sheet_state = "hidden"
    return new, n_modified


def _build_cumulative_detail_sheet(
    wb,
    prev_month: str,
    month: str,
    totals: dict[tuple[str, str, str], dict[str, float]],
    diffs: dict[tuple[str, str], dict[str, pd.Series]],
):
    src = wb[f"{prev_month} 누계(상세)"]
    new = wb.copy_worksheet(src)
    new.title = f"{month} 누계(상세)"
    new.sheet_state = "visible"
    a1 = new["A1"].value
    if isinstance(a1, str) and prev_month in a1:
        new["A1"] = a1.replace(prev_month, month)
    for hb, rowmap in [("본사", ROWS_HQ), ("법인", ROWS_CORP)]:
        for cat, r in rowmap.items():
            budget = totals[(cat, hb, "예산")]
            actual = totals[(cat, hb, "실적")]
            seg_diffs = diffs[(cat, hb)]
            for seg, (bcol, acol, ccol) in SEG_BLOCK_COLS.items():
                new[f"{bcol}{r}"] = round(budget[seg], 0)
                new[f"{acol}{r}"] = round(actual[seg], 0)
                total_diff = actual[seg] - budget[seg]
                if seg == "총합계" or abs(total_diff) >= COMMENT_THRESHOLD:
                    text = _format_comment(seg_diffs[seg], force_top1=(seg == "총합계"))
                    new[f"{ccol}{r}"] = text if text else None
                else:
                    new[f"{ccol}{r}"] = None
    src.sheet_state = "hidden"
    return new


def _reposition(wb, prev_month: str, month: str, suffix: str) -> None:
    name_new = f"{month} {suffix}"
    name_prev = f"{prev_month} {suffix}"
    idx_prev = wb.sheetnames.index(name_prev)
    idx_new = wb.sheetnames.index(name_new)
    wb.move_sheet(name_new, offset=(idx_prev + 1) - idx_new)


def _cross_validate(
    wb, sums_target: dict[int, dict[str, float]], month: str, months_all: list[str]
) -> tuple[float, float]:
    """1월~당월 실적 누계를 AC 시트 leaf-row 직접합으로 재계산 → 누계(상세) 총합계와 비교."""
    total = 0.0
    for m in months_all[:-1]:  # 1월~전월: 이미 리터럴 값으로 채워져 있음
        if f"{m}(AC)" not in wb.sheetnames:
            raise errors.SheetNotFound(
                f"교차검증 실패: '{m}(AC)' 시트가 템플릿에 없습니다. "
                "1월부터 전월까지의 개별월 시트가 모두 채워져 있어야 합니다."
            )
        ws = wb[f"{m}(AC)"]
        for r in LEAF_ROWS:
            for col in SEG_RAW_IDX:
                total += float(ws.cell(r, column_index_from_string(col)).value or 0)
    for r in LEAF_ROWS:  # 당월: 방금 집계한 sums 사용
        for col in SEG_RAW_IDX:
            total += sums_target.get(r, {}).get(col, 0.0)

    detail_ws = wb[f"{month} 누계(상세)"]
    grand_total_actual = sum(
        detail_ws[f"D{r}"].value or 0 for r in (list(ROWS_HQ.values()) + list(ROWS_CORP.values()))
    )  # D열 = 실적(총합계 블록)
    return total, float(grand_total_actual)


def _require_sheet(wb, name: str, hint: str = "") -> None:
    if name not in wb.sheetnames:
        suffix = f" {hint}" if hint else ""
        raise errors.SheetNotFound(
            f"시트를 찾을 수 없습니다: {name}. 사용 가능: {wb.sheetnames}.{suffix}"
        )


async def update_fc_month_report(
    raw_path: str,
    template_path: str,
    month: str,
    *,
    output_path: str | None = None,
    dry_run: bool = False,
    cross_validate_tolerance: float = 10.0,
    render_format: Literal["excel", "live_artifact", "both"] = "excel",
) -> FcMonthUpdateResult:
    """raw data + 전월 FC 보고서로 당월 배부판 자동 생성 (개별월(AC) 채우기 + 누계(AC)/누계(상세) 생성)."""
    n = _parse_month(month)
    prev_month = f"{n - 1}월"

    raw = excel_io.assert_xlsx_path(raw_path)
    tpl = excel_io.assert_xlsx_path(template_path)

    if output_path is not None:
        out = Path(output_path)
    else:
        out = tpl.parent / f"{tpl.stem}_{month}{tpl.suffix}"
    if out.resolve() == tpl.resolve():
        raise errors.OverwriteOriginalForbidden(
            "output_path가 원본 template_path와 동일합니다. 원본 직접 덮어쓰기는 허용되지 않습니다."
        )
    if not out.parent.exists():
        raise errors.WritePermissionDenied(f"출력 디렉터리가 없습니다: {out.parent}")

    wb = excel_io.load_workbook_safe(tpl, data_only=False)

    _require_sheet(
        wb,
        f"{month}(AC)",
        hint="대상월 스켈레톤 시트가 템플릿에 미리 존재해야 합니다.",
    )
    _require_sheet(wb, f"{prev_month}(AC)")
    for required in (f"{prev_month} 누계(AC)", f"{prev_month} 누계(상세)"):
        _require_sheet(wb, required)
    for forbidden in (f"{month} 누계(AC)", f"{month} 누계(상세)"):
        if forbidden in wb.sheetnames:
            raise errors.StructureMismatch(
                f"'{forbidden}' 시트가 이미 존재합니다. 이미 생성된 보고서를 입력한 것으로 보입니다."
            )

    sums, total_check, unmatched = _aggregate_org_actuals(str(raw), month)
    warnings_list: list[str] = []
    for org, amt in unmatched.items():
        if org != "DUMMY" and amt != 0:
            warnings_list.append(f"미매칭 조직 '{org}': {amt:,.0f}원 (집계에서 제외됨)")

    backup_path = None
    if not dry_run:
        backup_path = backup_mod.create_backup(tpl)

    ws_cur = wb[f"{month}(AC)"]
    ws_ref = wb[f"{prev_month}(AC)"]
    restored = _restore_ac_formulas(ws_cur, ws_ref, month, prev_month)
    written = _fill_ac_actuals(ws_cur, sums)
    ws_cur.sheet_state = "visible"

    months_all = _month_range(month)
    raw_df = _load_raw_df(str(raw), months_all)
    totals = _compute_totals(raw_df)
    diffs = _compute_account_diffs(raw_df)

    _, n_modified = _build_cumulative_ac_sheet(wb, prev_month, month)
    _build_cumulative_detail_sheet(wb, prev_month, month, totals, diffs)
    _reposition(wb, prev_month, month, "누계(AC)")
    _reposition(wb, prev_month, month, "누계(상세)")

    derived_cumulative_total, detail_grand_total = _cross_validate(wb, sums, month, months_all)
    raw_direct_total = (
        float(raw_df.loc[raw_df["gubun"] == "실적", "amount_krw"].sum()) if not raw_df.empty else 0.0
    )
    diff = max(
        abs(derived_cumulative_total - detail_grand_total),
        abs(raw_direct_total - detail_grand_total),
    )
    if diff > cross_validate_tolerance:
        raise errors.VerificationFailed(
            f"3-way 교차검증 오차({diff:,.2f}원)가 허용치({cross_validate_tolerance}원)를 초과했습니다. "
            f"raw_direct_total={raw_direct_total:,.2f}, "
            f"derived_cumulative_total={derived_cumulative_total:,.2f}, "
            f"detail_grand_total={detail_grand_total:,.2f}"
        )
    verification = FcMonthVerification(
        verified=True,
        raw_direct_total=raw_direct_total,
        derived_cumulative_total=derived_cumulative_total,
        detail_grand_total=detail_grand_total,
        diff=diff,
        tolerance=cross_validate_tolerance,
    )

    if not dry_run:
        wb.save(out)
        check_wb = load_workbook(out, data_only=False)
        try:
            if check_wb[f"{month}(AC)"].sheet_state != "visible":
                raise errors.VerificationFailed(f"저장 후 검증 실패: '{month}(AC)' 시트가 visible이 아닙니다.")
            for sheet_name in (f"{month} 누계(AC)", f"{month} 누계(상세)"):
                if sheet_name not in check_wb.sheetnames:
                    raise errors.VerificationFailed(f"저장 후 검증 실패: '{sheet_name}' 시트가 없습니다.")
            for sheet_name in (f"{prev_month} 누계(AC)", f"{prev_month} 누계(상세)"):
                if check_wb[sheet_name].sheet_state != "hidden":
                    raise errors.VerificationFailed(f"저장 후 검증 실패: '{sheet_name}' 시트가 hidden이 아닙니다.")
            sample_r, sample_col = LEAF_ROWS[0], next(iter(SEG_RAW_IDX))
            expected = round(sums.get(sample_r, {}).get(sample_col, 0.0), 2)
            actual_cell = check_wb[f"{month}(AC)"].cell(
                sample_r, column_index_from_string(sample_col)
            ).value
            if round(float(actual_cell or 0), 2) != expected:
                raise errors.VerificationFailed(
                    f"저장 후 검증 실패: 샘플 셀({sample_col}{sample_r}) 값 불일치 "
                    f"(기대={expected}, 실제={actual_cell})"
                )
        finally:
            check_wb.close()

    hints = ah.maybe_hints(
        render_format,
        artifact_type="verification_result",
        title=f"{month} 배부판 생성 결과",
    )

    return FcMonthUpdateResult(
        dry_run=dry_run,
        month=month,
        prev_month=prev_month,
        output_path=str(out) if not dry_run else None,
        backup_path=str(backup_path) if backup_path else None,
        ac_summary=FcMonthAcSummary(
            sheet=f"{month}(AC)",
            formulas_restored=restored,
            cells_written=written,
            unmatched_orgs={k: v for k, v in unmatched.items() if k != "DUMMY" and v != 0},
            raw_actual_total=total_check,
        ),
        cumulative_summary=FcMonthCumulativeSummary(
            ac_sheet_created=f"{month} 누계(AC)",
            detail_sheet_created=f"{month} 누계(상세)",
            prev_sheets_hidden=[f"{prev_month} 누계(AC)", f"{prev_month} 누계(상세)"],
            formula_chain_extended=n_modified,
        ),
        verification=verification,
        warnings=warnings_list,
        render_format=render_format,
        artifact_hints=hints,
    )
