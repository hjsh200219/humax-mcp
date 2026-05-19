"""update_allocation_rates tool — PRD §4.7."""

from __future__ import annotations

from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

from ..core import artifact_hints as ah
from ..core import backup as backup_mod
from ..core import errors, excel_io
from ..schemas import bp26
from ..schemas.requests import AllocationUpdate
from ..schemas.responses import (
    AllocationChange,
    AllocationUpdateData,
    AllocationUpdateResult,
)

RATE_KEYS = list(bp26.ALLOCATION_RATE_COLUMNS.values())
RATE_HEADER_KEYS = list(bp26.ALLOCATION_RATE_COLUMNS.keys())


def _validate_update(upd: AllocationUpdate, tolerance: float) -> None:
    for key, val in upd.new_rates.items():
        if not (0.0 <= float(val) <= 100.0):
            raise errors.InvalidRate(f"배부율은 0-100 범위여야 합니다. {key}={val}")
    rsum = sum(float(v) for v in upd.new_rates.values())
    if abs(rsum - 100.0) > tolerance:
        raise errors.RateSumNot100(
            f"배부율 합계가 100%가 아닙니다. cost_center={upd.cost_center}, 합계={rsum}%"
        )


async def update_allocation_rates(
    file_path: str,
    month: int,
    updates: list[AllocationUpdate] | list[dict],
    output_path: str,
    *,
    dry_run: bool = False,
    rate_tolerance: float = 0.01,
    render_format: Literal["excel", "live_artifact", "both"] = "excel",
) -> AllocationUpdateResult:
    """Update 26BP allocation rates with backup + dry-run + validation."""
    if not 1 <= month <= 12:
        raise errors.InvalidMonth(f"month는 1-12 범위여야 합니다. 입력: {month}")

    parsed: list[AllocationUpdate] = [
        u if isinstance(u, AllocationUpdate) else AllocationUpdate(**u) for u in updates
    ]
    for u in parsed:
        _validate_update(u, rate_tolerance)

    src = excel_io.assert_xlsx_path(file_path)
    out = Path(output_path)
    if out.resolve() == src.resolve():
        raise errors.OverwriteOriginalForbidden(
            "output_path가 원본 file_path와 동일합니다. 원본 직접 덮어쓰기는 허용되지 않습니다."
        )

    wb_check = excel_io.load_workbook_safe(src, data_only=True)
    if "예산+실적" not in wb_check.sheetnames:
        raise errors.SheetNotFound("예산+실적 시트를 찾을 수 없습니다.")
    headers_check = [c.value for c in wb_check["예산+실적"][1]]
    missing = [k for k in RATE_HEADER_KEYS if k not in headers_check]
    if missing:
        raise errors.SchemaMismatch(
            f"C30-C34 배부율 컬럼이 누락되었습니다. 스키마 v{bp26.SCHEMA_VERSION} 확인 필요. 누락: {missing}"
        )

    backup_path = None
    if not dry_run:
        backup_path = backup_mod.create_backup(src)

    wb = load_workbook(src)
    ws = wb["예산+실적"]
    headers = [c.value for c in ws[1]]
    cc_col_idx = headers.index("Cost Center") + 1 if "Cost Center" in headers else None
    basis_col_idx = headers.index("배부기준") + 1 if "배부기준" in headers else None
    rate_col_idxs = {k: headers.index(h) + 1 for h, k in bp26.ALLOCATION_RATE_COLUMNS.items()}

    if cc_col_idx is None or basis_col_idx is None:
        raise errors.SchemaMismatch("Cost Center 또는 배부기준 컬럼 미존재")

    changes: list[AllocationChange] = []
    for upd in parsed:
        target_rows: list[int] = []
        before_rates: dict[str, float] = {}
        for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            cc_val = row[cc_col_idx - 1]
            basis_val = row[basis_col_idx - 1]
            if str(cc_val) == upd.cost_center and str(basis_val or "") == upd.allocation_basis:
                target_rows.append(ridx)
                if not before_rates:
                    before_rates = {
                        k: float(row[rate_col_idxs[k] - 1] or 0) for k in RATE_KEYS
                    }
        if not target_rows:
            raise errors.CCBasisNotFound(
                f"cost_center={upd.cost_center}, allocation_basis={upd.allocation_basis} 일치 행을 찾을 수 없습니다."
            )
        if not dry_run:
            for ridx in target_rows:
                for k, v in upd.new_rates.items():
                    if k in rate_col_idxs:
                        ws.cell(row=ridx, column=rate_col_idxs[k]).value = float(v)
        changes.append(AllocationChange(
            cost_center=upd.cost_center,
            before=before_rates,
            after={k: float(upd.new_rates.get(k, 0.0)) for k in RATE_KEYS},
            rows_affected=len(target_rows),
        ))

    if not dry_run:
        wb.save(out)
        # Post-write verification — re-open and confirm rate cells match
        wb_check = load_workbook(out, data_only=True)
        ws_check = wb_check["예산+실적"]
        for upd in parsed:
            for ridx, row in enumerate(ws_check.iter_rows(min_row=2, values_only=True), start=2):
                cc_val = row[cc_col_idx - 1]
                basis_val = row[basis_col_idx - 1]
                if str(cc_val) == upd.cost_center and str(basis_val or "") == upd.allocation_basis:
                    for k, v in upd.new_rates.items():
                        actual = row[rate_col_idxs[k] - 1]
                        if actual is None or abs(float(actual) - float(v)) > 0.001:
                            raise errors.VerificationFailed(
                                f"쓰기 후 검증 실패: cc={upd.cost_center} {k}={actual} (expected {v}). 백업에서 복구하세요."
                            )

    data = AllocationUpdateData(
        output_path=str(out) if not dry_run else None,
        backup_path=str(backup_path) if backup_path else None,
        updates_applied=sum(c.rows_affected for c in changes),
        changes=changes,
    )

    hints = ah.maybe_hints(
        render_format,
        artifact_type="diff_cards",
        title=f"26.{month:02d} 배부율 변경 {'(dry-run)' if dry_run else '결과'}",
        preferred_chart="before_after_bar",
        comparison_columns=["before", "after"],
    )

    return AllocationUpdateResult(
        dry_run=dry_run,
        data=data,
        metadata={
            "month": month,
            "validation_violations": 0,
            "schema_version": bp26.SCHEMA_VERSION,
        },
        render_format=render_format,
        artifact_hints=hints,
    )
