"""US-009 generate_diff_candidates tests."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from humax_excel_mcp.core import errors
from humax_excel_mcp.tools.diff import generate_diff_candidates

pytestmark = pytest.mark.asyncio


async def test_no_candidates_when_files_equal(curr_month_path: Path, tmp_path: Path) -> None:
    twin = tmp_path / "twin.xlsx"
    twin.write_bytes(curr_month_path.read_bytes())
    res = await generate_diff_candidates(
        str(twin), str(curr_month_path), prev_sheet="예산+실적", curr_sheet="예산+실적"
    )
    assert res.success
    assert res.summary.candidates_found == 0


async def test_threshold_filtering(prev_month_path: Path, curr_month_path: Path) -> None:
    res = await generate_diff_candidates(
        str(prev_month_path),
        str(curr_month_path),
        prev_sheet="예산+실적",
        curr_sheet="예산+실적",
        threshold_million=0.001,
    )
    for c in res.candidates:
        assert abs(c.diff_million) >= 0.001


async def test_large_synthetic_diff_detected(
    curr_month_path: Path, tmp_path: Path
) -> None:
    prev = tmp_path / "prev_twin.xlsx"
    prev.write_bytes(curr_month_path.read_bytes())
    wb = load_workbook(curr_month_path)
    ws = wb["예산+실적"]
    headers = [c.value for c in ws[1]]
    aa_idx = headers.index("연간 실적") + 1
    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == "소조직":
            ws.cell(row=ridx, column=aa_idx).value = float(row[aa_idx - 1] or 0) + 500_000_000
            break
    wb.save(curr_month_path)
    res = await generate_diff_candidates(
        str(prev),
        str(curr_month_path),
        prev_sheet="예산+실적",
        curr_sheet="예산+실적",
        threshold_million=10.0,
    )
    assert res.summary.candidates_found >= 1
    top = res.candidates[0]
    assert top.severity == "HIGH"
    assert top.comment_draft and "백만" in top.comment_draft


async def test_comment_draft_format(curr_month_path: Path, tmp_path: Path) -> None:
    # First mark target row on both files (same gl), then bump only curr's actual
    wb = load_workbook(curr_month_path)
    ws = wb["예산+실적"]
    headers = [c.value for c in ws[1]]
    aa_idx = headers.index("연간 실적") + 1
    gl_idx = headers.index("G/L Account") + 1
    gl_name_idx = headers.index("G/L Account Name") + 1
    target_row = None
    base_actual = 0.0
    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == "소조직":
            target_row = ridx
            base_actual = float(row[aa_idx - 1] or 0)
            ws.cell(row=ridx, column=gl_idx).value = "511234"
            ws.cell(row=ridx, column=gl_name_idx).value = "급여"
            break
    wb.save(curr_month_path)
    assert target_row is not None

    prev = tmp_path / "prev_twin.xlsx"
    prev.write_bytes(curr_month_path.read_bytes())

    # Now bump curr's actual by 85M
    wb2 = load_workbook(curr_month_path)
    wb2["예산+실적"].cell(row=target_row, column=aa_idx).value = base_actual + 85_000_000
    wb2.save(curr_month_path)

    res = await generate_diff_candidates(
        str(prev),
        str(curr_month_path),
        prev_sheet="예산+실적",
        curr_sheet="예산+실적",
        include_comment_draft=True,
        threshold_million=10.0,
    )
    target = next((c for c in res.candidates if c.account_code == "511234"), None)
    assert target is not None
    assert target.comment_draft and "51" in target.comment_draft and "급여" in target.comment_draft


async def test_max_candidates_cap(prev_month_path: Path, curr_month_path: Path) -> None:
    res = await generate_diff_candidates(
        str(prev_month_path),
        str(curr_month_path),
        prev_sheet="예산+실적",
        curr_sheet="예산+실적",
        threshold_million=0.001,
        max_candidates=5,
    )
    assert res.summary.candidates_returned <= 5


async def test_prev_file_not_found(tmp_path: Path, curr_month_path: Path) -> None:
    with pytest.raises(errors.FileNotFound):
        await generate_diff_candidates(
            str(tmp_path / "missing.xlsx"),
            str(curr_month_path),
            prev_sheet="예산+실적",
            curr_sheet="예산+실적",
        )


async def test_live_artifact_hints(prev_month_path: Path, curr_month_path: Path) -> None:
    res = await generate_diff_candidates(
        str(prev_month_path),
        str(curr_month_path),
        prev_sheet="예산+실적",
        curr_sheet="예산+실적",
        render_format="live_artifact",
    )
    assert res.artifact_hints is not None
    assert res.artifact_hints.type == "diff_cards"
