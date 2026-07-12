"""US-S3 workbook_cache tests — 캐시 히트 / mtime 무효화 / copy 격리 / LRU."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from humax_excel_mcp.core import errors, excel_io, workbook_cache


@pytest.fixture(autouse=True)
def _fresh_cache():
    workbook_cache.clear()
    yield
    workbook_cache.clear()


def test_second_call_hits_cache(sample_26bp_path: Path, monkeypatch) -> None:
    calls = {"n": 0}
    orig = excel_io.load_workbook_safe

    def counting(path, **kw):
        calls["n"] += 1
        return orig(path, **kw)

    monkeypatch.setattr(excel_io, "load_workbook_safe", counting)
    df1 = workbook_cache.get_dataframe(str(sample_26bp_path), "예산+실적")
    df2 = workbook_cache.get_dataframe(str(sample_26bp_path), "예산+실적")
    assert calls["n"] == 1
    assert workbook_cache.stats["hits"] == 1
    assert df1.equals(df2)


def test_file_modification_invalidates(sample_26bp_path: Path) -> None:
    workbook_cache.get_dataframe(str(sample_26bp_path), "예산+실적")
    wb = load_workbook(sample_26bp_path)
    wb["예산+실적"].cell(row=2, column=2).value = "HUS"
    wb.save(sample_26bp_path)
    df = workbook_cache.get_dataframe(str(sample_26bp_path), "예산+실적")
    assert workbook_cache.stats["misses"] == 2
    assert df.iloc[0]["company"] == "HUS"


def test_returned_df_is_isolated_copy(sample_26bp_path: Path) -> None:
    df1 = workbook_cache.get_dataframe(str(sample_26bp_path), "예산+실적")
    df1["division"] = "oops"
    df2 = workbook_cache.get_dataframe(str(sample_26bp_path), "예산+실적")
    assert (df2["division"] != "oops").all()
    assert workbook_cache.stats["hits"] == 1


def test_invalidate_removes_entries(sample_26bp_path: Path) -> None:
    workbook_cache.get_dataframe(str(sample_26bp_path), "예산+실적")
    workbook_cache.invalidate(sample_26bp_path)
    workbook_cache.get_dataframe(str(sample_26bp_path), "예산+실적")
    assert workbook_cache.stats["misses"] == 2
    assert workbook_cache.stats["hits"] == 0


def test_lru_eviction(sample_26bp_path: Path, empty_xlsx: Path, monkeypatch) -> None:
    monkeypatch.setattr(workbook_cache, "_MAX_ENTRIES", 1)
    workbook_cache.get_dataframe(str(sample_26bp_path), "예산+실적")
    workbook_cache.get_dataframe(str(empty_xlsx), "예산+실적")
    workbook_cache.get_dataframe(str(sample_26bp_path), "예산+실적")
    assert workbook_cache.stats["misses"] == 3
    assert workbook_cache.stats["hits"] == 0


def test_missing_file_raises(tmp_path: Path) -> None:
    with pytest.raises(errors.FileNotFound):
        workbook_cache.get_dataframe(str(tmp_path / "nope.xlsx"), "예산+실적")


def test_missing_sheet_raises(sample_26bp_path: Path) -> None:
    with pytest.raises(errors.SheetNotFound):
        workbook_cache.get_dataframe(str(sample_26bp_path), "없는시트")
