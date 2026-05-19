"""US-016 apply_golden_template tests."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from humax_excel_mcp.core import errors
from humax_excel_mcp.tools.template_engine import apply_golden_template

pytestmark = pytest.mark.asyncio

ROOT = Path(__file__).resolve().parents[2]
FIX_DIR = ROOT / "fixtures" / "templates"


def _template_path(template_type: str) -> Path:
    return FIX_DIR / f"{template_type}.xlsx"


@pytest.mark.parametrize("template_type", ["humax_allocation", "humax_account", "evcs_account"])
async def test_happy_path_per_template_type(
    template_type: str, sample_26bp_path: Path, tmp_path: Path
) -> None:
    out = tmp_path / f"{template_type}_out.xlsx"
    res = await apply_golden_template(
        source_file=str(sample_26bp_path),
        template_path=str(_template_path(template_type)),
        template_type=template_type,
        output_path=str(out),
        month=3,
    )
    assert res.success
    assert res.template_type == template_type
    assert out.exists()
    assert res.backup_path
    assert len(res.sheets_processed) >= 1


async def test_dry_run_no_file_write(
    sample_26bp_path: Path, tmp_path: Path
) -> None:
    out = tmp_path / "out.xlsx"
    res = await apply_golden_template(
        source_file=str(sample_26bp_path),
        template_path=str(_template_path("humax_account")),
        template_type="humax_account",
        output_path=str(out),
        month=3,
        dry_run=True,
    )
    assert res.dry_run is True
    assert res.output_path is None
    assert res.backup_path is None
    assert not out.exists()


async def test_overwrite_original_forbidden(sample_26bp_path: Path) -> None:
    with pytest.raises(errors.OverwriteOriginalForbidden):
        await apply_golden_template(
            source_file=str(sample_26bp_path),
            template_path=str(_template_path("humax_account")),
            template_type="humax_account",
            output_path=str(sample_26bp_path),
            month=3,
        )


async def test_template_not_found(sample_26bp_path: Path, tmp_path: Path) -> None:
    with pytest.raises(errors.TemplateNotFound):
        await apply_golden_template(
            source_file=str(sample_26bp_path),
            template_path=str(tmp_path / "missing.xlsx"),
            template_type="humax_account",
            output_path=str(tmp_path / "out.xlsx"),
            month=3,
        )


async def test_binding_not_found_for_unknown_type(
    sample_26bp_path: Path, tmp_path: Path
) -> None:
    with pytest.raises((errors.BindingNotFound, ValueError, TypeError)):
        await apply_golden_template(
            source_file=str(sample_26bp_path),
            template_path=str(_template_path("humax_account")),
            template_type="zzz_unknown",  # type: ignore[arg-type]
            output_path=str(tmp_path / "out.xlsx"),
            month=3,
        )


async def test_template_malformed_when_sidecar_missing(
    sample_26bp_path: Path, tmp_path: Path
) -> None:
    # Copy xlsx without sidecar
    bad_template = tmp_path / "humax_account.xlsx"
    bad_template.write_bytes(_template_path("humax_account").read_bytes())
    with pytest.raises(errors.TemplateMalformed):
        await apply_golden_template(
            source_file=str(sample_26bp_path),
            template_path=str(bad_template),
            template_type="humax_account",
            output_path=str(tmp_path / "out.xlsx"),
            month=3,
        )


async def test_merged_cells_preserved(
    sample_26bp_path: Path, tmp_path: Path
) -> None:
    out = tmp_path / "out.xlsx"
    # Reference has merged cells; assert output preserves them (>0, not just >=0)
    src_template_merged = len(
        load_workbook(_template_path("humax_allocation"))["3월 누계"].merged_cells.ranges
    )
    res = await apply_golden_template(
        source_file=str(sample_26bp_path),
        template_path=str(_template_path("humax_allocation")),
        template_type="humax_allocation",
        output_path=str(out),
        month=3,
    )
    assert res.success
    wb = load_workbook(out)
    ws = wb["3월 누계"]
    assert len(ws.merged_cells.ranges) == src_template_merged


async def test_formula_cells_not_overwritten(
    sample_26bp_path: Path, tmp_path: Path
) -> None:
    out = tmp_path / "out.xlsx"
    res = await apply_golden_template(
        source_file=str(sample_26bp_path),
        template_path=str(_template_path("humax_account")),
        template_type="humax_account",
        output_path=str(out),
        month=3,
    )
    sheet_summary = next(s for s in res.sheets_processed if s.sheet_name == "요약")
    # 요약 sheet has formulas (sidecar shows formula count). Assert at least one preserved.
    assert sheet_summary.formulas_preserved > 0


async def test_multi_sheet_binding_all_processed(
    sample_26bp_path: Path, tmp_path: Path
) -> None:
    out = tmp_path / "out.xlsx"
    res = await apply_golden_template(
        source_file=str(sample_26bp_path),
        template_path=str(_template_path("humax_allocation")),
        template_type="humax_allocation",
        output_path=str(out),
        month=3,
    )
    # binding currently has 1 sheet (3월 누계) but ensure all bound sheets processed
    bound_sheets = {s.sheet_name for s in res.sheets_processed}
    assert "3월 누계" in bound_sheets


async def test_invalid_month(sample_26bp_path: Path, tmp_path: Path) -> None:
    with pytest.raises(errors.InvalidMonth):
        await apply_golden_template(
            source_file=str(sample_26bp_path),
            template_path=str(_template_path("humax_account")),
            template_type="humax_account",
            output_path=str(tmp_path / "out.xlsx"),
            month=13,
        )


@pytest.mark.asyncio
async def test_humax_account_filter_matches_valid_humax_companies() -> None:
    """US-025 AC5: humax_account filter_values matches raw_bp26.VALID_HUMAX_COMPANIES."""
    from humax_excel_mcp.core.template_bindings import HUMAX_ACCOUNT_BINDING
    from humax_excel_mcp.schemas import raw_bp26

    sheet = HUMAX_ACCOUNT_BINDING.sheets[0]
    assert sheet.row_selection.filter_values == raw_bp26.VALID_HUMAX_COMPANIES
    assert sheet.row_selection.filter_values == ["HKR", "HMX", "HUS", "HUK", "HBR", "HSZ"]


@pytest.mark.asyncio
async def test_aggregated_from_raw_flow_through_bindings(tmp_path: Path) -> None:
    """US-025 AC3: aggregator output flows through existing bindings."""
    import pandas as pd

    from humax_excel_mcp.core.aggregator import aggregate_to_bp26
    from humax_excel_mcp.core.template_bindings import HUMAX_ACCOUNT_BINDING

    # Tiny raw df: HKR, one CC, one GL
    raw = pd.DataFrame([{
        "division_type": "실적",
        "year": "26년",
        "month": "1월",
        "company": "HKR",
        "cost_center": 101,
        "cost_center_name": "CC101",
        "gl_account": 510000,
        "gl_account_name": "GL510000",
        "org_l1": "사업그룹",
        "amount_krw": 1000.0,
        "rate_evcs_domestic": 0.0,
        "rate_evcs_overseas": 0.0,
    }])
    result = aggregate_to_bp26(raw, target_month=3, expand_evcs=False)
    binding_sheet = HUMAX_ACCOUNT_BINDING.sheets[0]
    # All column_map source keys must be present in aggregator output
    for source_key in binding_sheet.column_map.values():
        assert source_key in result.df.columns, f"Missing key in aggregator output: {source_key}"
    # HKR present (6-company filter accepts it)
    assert "HKR" in result.df["company"].values
