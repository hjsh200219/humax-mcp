"""US-011 update_allocation_rates tests."""

from __future__ import annotations

import hashlib
from pathlib import Path

import pytest
from openpyxl import load_workbook

from humax_excel_mcp.core import errors
from humax_excel_mcp.schemas.requests import AllocationUpdate
from humax_excel_mcp.tools.allocation_get import get_allocation_rates
from humax_excel_mcp.tools.allocation_set import update_allocation_rates

pytestmark = pytest.mark.asyncio


def _sha(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


async def _pick_update(sample_26bp_path: Path) -> AllocationUpdate:
    res = await get_allocation_rates(str(sample_26bp_path), month=3)
    row = res.data[0]
    return AllocationUpdate(
        cost_center=row.cost_center,
        allocation_basis=row.allocation_basis,
        new_rates={"STB": 40.0, "Mobility": 10.0, "EVCS_domestic": 25.0, "EVCS_overseas": 25.0},
    )


async def test_dry_run_preserves_file(sample_26bp_path: Path, tmp_path: Path) -> None:
    upd = await _pick_update(sample_26bp_path)
    pre = _sha(sample_26bp_path)
    res = await update_allocation_rates(
        str(sample_26bp_path),
        month=3,
        updates=[upd],
        output_path=str(tmp_path / "out.xlsx"),
        dry_run=True,
    )
    assert res.dry_run is True
    assert res.data.output_path is None
    assert res.data.backup_path is None
    assert _sha(sample_26bp_path) == pre
    assert len(res.data.changes) == 1


async def test_apply_changes_actual(sample_26bp_path: Path, tmp_path: Path) -> None:
    upd = await _pick_update(sample_26bp_path)
    out = tmp_path / "out.xlsx"
    res = await update_allocation_rates(
        str(sample_26bp_path),
        month=3,
        updates=[upd],
        output_path=str(out),
    )
    assert res.dry_run is False
    assert out.exists()
    assert res.data.backup_path
    assert Path(res.data.backup_path).exists()
    assert res.data.updates_applied >= 1


async def test_overwrite_original_forbidden(sample_26bp_path: Path) -> None:
    upd = await _pick_update(sample_26bp_path)
    with pytest.raises(errors.OverwriteOriginalForbidden):
        await update_allocation_rates(
            str(sample_26bp_path),
            month=3,
            updates=[upd],
            output_path=str(sample_26bp_path),
        )


async def test_rate_sum_not_100(sample_26bp_path: Path, tmp_path: Path) -> None:
    bad = AllocationUpdate(
        cost_center="100001",
        allocation_basis="경영지원부문",
        new_rates={"STB": 30.0, "Mobility": 30.0, "EVCS_domestic": 30.0, "EVCS_overseas": 30.0},
    )
    with pytest.raises(errors.RateSumNot100):
        await update_allocation_rates(
            str(sample_26bp_path),
            month=3,
            updates=[bad],
            output_path=str(tmp_path / "out.xlsx"),
        )


async def test_invalid_rate_range(sample_26bp_path: Path, tmp_path: Path) -> None:
    bad = AllocationUpdate(
        cost_center="100001",
        allocation_basis="x",
        new_rates={"STB": 150.0, "Mobility": -50.0, "EVCS_domestic": 0.0, "EVCS_overseas": 0.0},
    )
    with pytest.raises(errors.InvalidRate):
        await update_allocation_rates(
            str(sample_26bp_path),
            month=3,
            updates=[bad],
            output_path=str(tmp_path / "out.xlsx"),
        )


async def test_cc_basis_not_found(sample_26bp_path: Path, tmp_path: Path) -> None:
    bad = AllocationUpdate(
        cost_center="ZZZ-NEVER",
        allocation_basis="NOPE",
        new_rates={"STB": 25.0, "Mobility": 25.0, "EVCS_domestic": 25.0, "EVCS_overseas": 25.0},
    )
    with pytest.raises(errors.CCBasisNotFound):
        await update_allocation_rates(
            str(sample_26bp_path),
            month=3,
            updates=[bad],
            output_path=str(tmp_path / "out.xlsx"),
        )


async def test_invalid_month(sample_26bp_path: Path, tmp_path: Path) -> None:
    upd = AllocationUpdate(
        cost_center="x", allocation_basis="x",
        new_rates={"STB": 25.0, "Mobility": 25.0, "EVCS_domestic": 25.0, "EVCS_overseas": 25.0},
    )
    with pytest.raises(errors.InvalidMonth):
        await update_allocation_rates(
            str(sample_26bp_path), month=0, updates=[upd], output_path=str(tmp_path / "out.xlsx")
        )


async def test_tolerance_allows_drift(sample_26bp_path: Path, tmp_path: Path) -> None:
    upd = await _pick_update(sample_26bp_path)
    drift = AllocationUpdate(
        cost_center=upd.cost_center,
        allocation_basis=upd.allocation_basis,
        new_rates={"STB": 33.33, "Mobility": 33.33, "EVCS_domestic": 33.34, "EVCS_overseas": 0.0},
    )
    res = await update_allocation_rates(
        str(sample_26bp_path),
        month=3,
        updates=[drift],
        output_path=str(tmp_path / "out.xlsx"),
        dry_run=True,
        rate_tolerance=0.01,
    )
    assert res.success


async def test_live_artifact_hints(sample_26bp_path: Path, tmp_path: Path) -> None:
    upd = await _pick_update(sample_26bp_path)
    res = await update_allocation_rates(
        str(sample_26bp_path),
        month=3,
        updates=[upd],
        output_path=str(tmp_path / "out.xlsx"),
        dry_run=True,
        render_format="live_artifact",
    )
    assert res.artifact_hints is not None
    assert res.artifact_hints.type == "diff_cards"
    assert res.artifact_hints.preferred_chart == "before_after_bar"


async def test_only_rate_cells_modified(sample_26bp_path: Path, tmp_path: Path) -> None:
    upd = await _pick_update(sample_26bp_path)
    out = tmp_path / "out.xlsx"
    pre_wb = load_workbook(sample_26bp_path, data_only=True)
    pre_ws = pre_wb["예산+실적"]
    pre_a3 = pre_ws["A3"].value
    pre_b3 = pre_ws["B3"].value
    await update_allocation_rates(
        str(sample_26bp_path), month=3, updates=[upd], output_path=str(out)
    )
    wb = load_workbook(out, data_only=True)
    ws = wb["예산+실적"]
    assert ws["A3"].value == pre_a3
    assert ws["B3"].value == pre_b3
