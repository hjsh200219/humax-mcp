"""US-010 get_allocation_rates tests."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from humax_excel_mcp.core import errors
from humax_excel_mcp.tools.allocation_get import get_allocation_rates

pytestmark = pytest.mark.asyncio


async def test_basic_get(sample_26bp_path: Path) -> None:
    res = await get_allocation_rates(str(sample_26bp_path), month=3)
    assert res.success
    assert res.metadata["month"] == 3
    assert res.metadata["unique_rates_count"] >= 1
    assert res.metadata["rate_sum_violations"] == 0


async def test_rate_sum_100_default_fixture(sample_26bp_path: Path) -> None:
    res = await get_allocation_rates(str(sample_26bp_path), month=3)
    for row in res.data:
        assert row.rate_sum_ok is True
        assert abs(row.rate_sum - 100.0) < 0.01


async def test_invalid_month(sample_26bp_path: Path) -> None:
    with pytest.raises(errors.InvalidMonth):
        await get_allocation_rates(str(sample_26bp_path), month=0)
    with pytest.raises(errors.InvalidMonth):
        await get_allocation_rates(str(sample_26bp_path), month=13)


async def test_invalid_company(sample_26bp_path: Path) -> None:
    with pytest.raises(errors.InvalidCompany):
        await get_allocation_rates(str(sample_26bp_path), month=3, company="ZZZ")


async def test_file_not_found(tmp_path: Path) -> None:
    with pytest.raises(errors.FileNotFound):
        await get_allocation_rates(str(tmp_path / "missing.xlsx"), month=3)


async def test_rate_sum_violation_counted(sample_26bp_path: Path) -> None:
    wb = load_workbook(sample_26bp_path)
    ws = wb["예산+실적"]
    headers = [c.value for c in ws[1]]
    stb_idx = headers.index("STB 배부율") + 1
    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == "소조직":
            ws.cell(row=ridx, column=stb_idx).value = 999.0
            break
    wb.save(sample_26bp_path)
    res = await get_allocation_rates(str(sample_26bp_path), month=3)
    assert res.metadata["rate_sum_violations"] >= 1


async def test_live_artifact_hints(sample_26bp_path: Path) -> None:
    res = await get_allocation_rates(
        str(sample_26bp_path), month=3, render_format="live_artifact"
    )
    assert res.artifact_hints is not None
    assert res.artifact_hints.type == "table_with_chart"
    assert res.artifact_hints.preferred_chart == "stacked_bar"


async def test_filter_company(sample_26bp_path: Path) -> None:
    res = await get_allocation_rates(str(sample_26bp_path), month=3, company="HMX")
    assert res.metadata["filter_company"] == "HMX"


async def test_schema_mismatch(tmp_path: Path) -> None:
    from openpyxl import Workbook
    p = tmp_path / "bad.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "예산+실적"
    ws.append(["구분", "Company"])
    wb.save(p)
    with pytest.raises(errors.SchemaMismatch):
        await get_allocation_rates(str(p), month=3)
