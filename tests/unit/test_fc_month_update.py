"""US-023 update_fc_month_report tests."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from humax_excel_mcp.core import errors
from humax_excel_mcp.core.fc_report_layout import ALIAS_TO_ROW, LEAF_ROWS, SEG_RAW_IDX
from humax_excel_mcp.tools.fc_month_update import (
    RAW_SHEET_NAME,
    _aggregate_org_actuals,
    _fill_ac_actuals,
    _format_comment,
    _month_range,
    _parse_month,
    _restore_ac_formulas,
    update_fc_month_report,
)

pytestmark = pytest.mark.asyncio

SEG_COLS_BY_NAME = {  # SEG_RAW_IDX 중 SEG_COLS_RAW와 공유되는 STB~건물 6개 열
    "STB": ("J", 45), "Mobility": ("K", 46), "EVCS_in": ("L", 47),
    "EVCS_out": ("M", 48), "공통": ("N", 49), "건물": ("O", 50),
}
RAW_NCOLS = 62  # idx61(Shared)까지 필요


def _raw_row(gubun: str, month: str, company: str, org: str, daegyejeong_re: str, segs: dict[str, float]) -> list:
    row = [None] * RAW_NCOLS
    row[0] = gubun
    row[2] = month
    row[5] = company
    row[10] = org
    row[17] = daegyejeong_re
    for name, (_letter, idx) in SEG_COLS_BY_NAME.items():
        row[idx] = segs.get(name, 0)
    row[23] = sum(segs.values())  # amount_krw == 세그먼트 합 (Q~W, Shared는 0으로 둠)
    for idx in range(51, 58):  # Q~W
        row[idx] = 0
    row[61] = 0  # Shared
    return row


def _build_raw_workbook(path: Path, rows: list[list]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = RAW_SHEET_NAME
    for _ in range(3):  # 데이터는 4행부터 (min_row=4)
        ws.append([None] * RAW_NCOLS)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


# ---- 순수 헬퍼 함수 단위 테스트 ----------------------------------------------


async def test_parse_month_valid():
    assert _parse_month("5월") == 5


@pytest.mark.parametrize("bad", ["1월", "13월", "0월", "May", "5", "5월차"])
async def test_parse_month_invalid(bad):
    with pytest.raises(errors.InvalidMonth):
        _parse_month(bad)


async def test_month_range():
    assert _month_range("3월") == ["1월", "2월", "3월"]


async def test_aggregate_org_actuals_sums_matched_orgs(tmp_path: Path):
    rows = [
        _raw_row("실적", "5월", "HKR", "HUS", "11 급여", {"STB": 100.0, "Mobility": 50.0}),
        _raw_row("실적", "5월", "HMX", "HMX", "11 급여", {"STB": 10.0}),
        _raw_row("실적", "5월", "HKR", "DUMMY", "11 급여", {"STB": 0.0}),
        _raw_row("예산", "5월", "HKR", "HUS", "11 급여", {"STB": 999.0}),  # 예산은 제외
        _raw_row("실적", "4월", "HKR", "HUS", "11 급여", {"STB": 777.0}),  # 다른 월은 제외
    ]
    raw_path = _build_raw_workbook(tmp_path / "raw.xlsx", rows)

    sums, total_check, unmatched = _aggregate_org_actuals(str(raw_path), "5월")

    hus_row = ALIAS_TO_ROW["HUS"]
    hmx_row = ALIAS_TO_ROW["HMX"]
    assert sums[hus_row]["J"] == 100.0
    assert sums[hus_row]["K"] == 50.0
    assert sums[hmx_row]["J"] == 10.0
    assert total_check == pytest.approx(160.0)
    assert unmatched == {"DUMMY": 0.0}  # DUMMY는 항상 0원이지만 defaultdict라 키 자체는 생성됨


async def test_aggregate_org_actuals_unmatched_org_nonzero(tmp_path: Path):
    rows = [_raw_row("실적", "5월", "HKR", "신규조직", "11 급여", {"STB": 42.0})]
    raw_path = _build_raw_workbook(tmp_path / "raw.xlsx", rows)

    sums, total_check, unmatched = _aggregate_org_actuals(str(raw_path), "5월")

    assert unmatched == {"신규조직": 42.0}
    assert total_check == pytest.approx(42.0)


async def test_aggregate_org_actuals_empty_result(tmp_path: Path):
    rows = [_raw_row("예산", "5월", "HKR", "HUS", "11 급여", {"STB": 1.0})]
    raw_path = _build_raw_workbook(tmp_path / "raw.xlsx", rows)

    with pytest.raises(errors.EmptyResult):
        _aggregate_org_actuals(str(raw_path), "5월")


async def test_restore_ac_formulas_copies_and_replaces_month():
    wb = Workbook()
    ws_ref = wb.active
    ws_ref.title = "4월(AC)"
    ws_cur = wb.create_sheet("5월(AC)")
    ws_ref.cell(2, 2).value = "='4월(BP)'!B2"
    ws_cur.cell(2, 2).value = None  # 빈 셀이라 복원 대상

    restored = _restore_ac_formulas(ws_cur, ws_ref, "5월", "4월")

    assert restored == 1
    assert ws_cur.cell(2, 2).value == "='5월(BP)'!B2"


async def test_restore_ac_formulas_skips_already_filled_cells():
    wb = Workbook()
    ws_ref = wb.active
    ws_ref.title = "4월(AC)"
    ws_cur = wb.create_sheet("5월(AC)")
    ws_ref.cell(2, 2).value = "='4월(BP)'!B2"
    ws_cur.cell(2, 2).value = 123  # 이미 값이 있으므로 건드리지 않음

    restored = _restore_ac_formulas(ws_cur, ws_ref, "5월", "4월")

    assert restored == 0
    assert ws_cur.cell(2, 2).value == 123


async def test_fill_ac_actuals_writes_all_leaf_cells():
    wb = Workbook()
    ws = wb.active
    sums = {ALIAS_TO_ROW["HUS"]: {"J": 1.5, "K": 2.5}}

    written = _fill_ac_actuals(ws, sums)

    assert written == len(LEAF_ROWS) * len(SEG_RAW_IDX)
    from openpyxl.utils import column_index_from_string

    assert ws.cell(ALIAS_TO_ROW["HUS"], column_index_from_string("J")).value == 1.5
    assert ws.cell(ALIAS_TO_ROW["HMX"], column_index_from_string("J")).value == 0.0


async def test_format_comment_orders_by_magnitude_and_caps_topn():
    import pandas as pd

    diffs = pd.Series({
        "11 급여": 10_000_000.0,
        "53 4대보험료": -6_000_000.0,
        "13 퇴직급여": 5_500_000.0,
        "16 여비교통비": 1_000_000.0,  # COMMENT_ITEM_MIN 미달 -> 제외
    })

    text = _format_comment(diffs)

    assert text == "11 급여 +10백만 / 53 4대보험료 △6백만 / 13 퇴직급여 +6백만"


async def test_format_comment_force_top1_when_below_threshold():
    import pandas as pd

    diffs = pd.Series({"11 급여": 1_000_000.0})

    assert _format_comment(diffs, force_top1=False) == ""
    assert _format_comment(diffs, force_top1=True) == "11 급여 +1백만"


async def test_format_comment_empty_series_returns_empty_string():
    import pandas as pd

    assert _format_comment(pd.Series(dtype=float)) == ""


# ---- 오케스트레이션: 사전조건 에러 ---------------------------------------------


def _minimal_template(path: Path, *, prev_month: str, month: str, include_month_skeleton: bool = True,
                       pre_existing_cumulative: bool = False) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{prev_month}(AC)"
    if include_month_skeleton:
        cur = wb.create_sheet(f"{month}(AC)")
        cur.sheet_state = "hidden"
    ac_cum = wb.create_sheet(f"{prev_month} 누계(AC)")
    ac_cum["A1"] = f"{prev_month} 누계(AC)"
    detail_cum = wb.create_sheet(f"{prev_month} 누계(상세)")
    detail_cum["A1"] = f"{prev_month} 누계(상세)"
    if pre_existing_cumulative:
        wb.create_sheet(f"{month} 누계(AC)")
    wb.save(path)
    return path


def _minimal_raw(path: Path, *, month: str = "5월", has_actual_row: bool = True) -> Path:
    rows = []
    if has_actual_row:
        rows.append(_raw_row("실적", month, "HKR", "HUS", "11 급여", {"STB": 100.0}))
    return _build_raw_workbook(path, rows)


async def test_missing_raw_file_raises_file_not_found(tmp_path: Path):
    template = _minimal_template(tmp_path / "tpl.xlsx", prev_month="4월", month="5월")
    with pytest.raises(errors.FileNotFound):
        await update_fc_month_report(
            raw_path=str(tmp_path / "missing.xlsx"),
            template_path=str(template),
            month="5월",
        )


async def test_invalid_month_format_raises(tmp_path: Path):
    raw = _minimal_raw(tmp_path / "raw.xlsx")
    template = _minimal_template(tmp_path / "tpl.xlsx", prev_month="4월", month="5월")
    with pytest.raises(errors.InvalidMonth):
        await update_fc_month_report(raw_path=str(raw), template_path=str(template), month="1월")


async def test_overwrite_original_forbidden(tmp_path: Path):
    raw = _minimal_raw(tmp_path / "raw.xlsx")
    template = _minimal_template(tmp_path / "tpl.xlsx", prev_month="4월", month="5월")
    with pytest.raises(errors.OverwriteOriginalForbidden):
        await update_fc_month_report(
            raw_path=str(raw), template_path=str(template), month="5월", output_path=str(template)
        )


async def test_write_permission_denied_when_output_dir_missing(tmp_path: Path):
    raw = _minimal_raw(tmp_path / "raw.xlsx")
    template = _minimal_template(tmp_path / "tpl.xlsx", prev_month="4월", month="5월")
    bad_out = tmp_path / "no_such_dir" / "out.xlsx"
    with pytest.raises(errors.WritePermissionDenied):
        await update_fc_month_report(
            raw_path=str(raw), template_path=str(template), month="5월", output_path=str(bad_out)
        )


async def test_missing_month_skeleton_sheet_raises_sheet_not_found(tmp_path: Path):
    raw = _minimal_raw(tmp_path / "raw.xlsx")
    template = _minimal_template(
        tmp_path / "tpl.xlsx", prev_month="4월", month="5월", include_month_skeleton=False
    )
    with pytest.raises(errors.SheetNotFound):
        await update_fc_month_report(raw_path=str(raw), template_path=str(template), month="5월")


async def test_structure_mismatch_when_cumulative_sheet_already_exists(tmp_path: Path):
    raw = _minimal_raw(tmp_path / "raw.xlsx")
    template = _minimal_template(
        tmp_path / "tpl.xlsx", prev_month="4월", month="5월", pre_existing_cumulative=True
    )
    with pytest.raises(errors.StructureMismatch):
        await update_fc_month_report(raw_path=str(raw), template_path=str(template), month="5월")


async def test_empty_result_when_raw_has_no_matching_rows(tmp_path: Path):
    raw = _minimal_raw(tmp_path / "raw.xlsx", month="5월", has_actual_row=False)
    template = _minimal_template(tmp_path / "tpl.xlsx", prev_month="4월", month="5월")
    with pytest.raises(errors.EmptyResult):
        await update_fc_month_report(raw_path=str(raw), template_path=str(template), month="5월")


# ---- 오케스트레이션: happy path (자기 일관적 fixture) ---------------------------


def _build_happy_path_fixtures(tmp_path: Path) -> tuple[Path, Path]:
    """1월/2월 raw 실적 데이터를 만들고, 1월(AC)는 같은 집계 로직으로 직접 채워
    cross-validate가 수학적으로 0에 일치하도록 구성한다 (Q~W/Shared는 항상 0)."""
    jan_rows = [
        _raw_row("실적", "1월", "HKR", "HUS", "11 급여",
                 {"STB": 1_000_000, "Mobility": 2_000_000, "EVCS_in": 500_000,
                  "EVCS_out": 300_000, "공통": 200_000, "건물": 100_000}),
        _raw_row("실적", "1월", "HMX", "HMX", "11 급여",
                 {"STB": 500_000, "Mobility": 400_000, "EVCS_in": 300_000,
                  "EVCS_out": 200_000, "공통": 100_000, "건물": 50_000}),
    ]
    feb_rows = [
        _raw_row("실적", "2월", "HKR", "HUS", "11 급여",
                 {"STB": 1_100_000, "Mobility": 2_100_000, "EVCS_in": 510_000,
                  "EVCS_out": 310_000, "공통": 210_000, "건물": 110_000}),
        _raw_row("실적", "2월", "HMX", "HMX", "11 급여",
                 {"STB": 520_000, "Mobility": 410_000, "EVCS_in": 310_000,
                  "EVCS_out": 210_000, "공통": 110_000, "건물": 60_000}),
    ]
    raw_path = _build_raw_workbook(tmp_path / "raw.xlsx", jan_rows + feb_rows)

    template_path = tmp_path / "template.xlsx"
    wb = Workbook()
    jan_ac = wb.active
    jan_ac.title = "1월(AC)"
    jan_ac.cell(2, 2).value = "='1월(BP)'!B2"  # _restore_ac_formulas 커버용 더미 수식

    feb_ac = wb.create_sheet("2월(AC)")
    feb_ac.sheet_state = "hidden"

    jan_sums, _, _ = _aggregate_org_actuals(str(raw_path), "1월")
    _fill_ac_actuals(jan_ac, jan_sums)

    jan_cum_ac = wb.create_sheet("1월 누계(AC)")
    jan_cum_ac["A1"] = "1월 누계(AC)"
    jan_cum_ac["C5"] = "='1월(AC)'!C5"

    jan_cum_detail = wb.create_sheet("1월 누계(상세)")
    jan_cum_detail["A1"] = "1월 누계(상세)"

    wb.save(template_path)
    return raw_path, template_path


async def test_happy_path_full_orchestration(tmp_path: Path):
    raw_path, template_path = _build_happy_path_fixtures(tmp_path)
    out_path = tmp_path / "out.xlsx"

    result = await update_fc_month_report(
        raw_path=str(raw_path),
        template_path=str(template_path),
        month="2월",
        output_path=str(out_path),
    )

    assert result.success
    assert result.dry_run is False
    assert result.month == "2월"
    assert result.prev_month == "1월"
    assert result.output_path == str(out_path)
    assert result.backup_path is not None
    assert out_path.exists()

    assert result.ac_summary.sheet == "2월(AC)"
    assert result.ac_summary.cells_written == len(LEAF_ROWS) * len(SEG_RAW_IDX)
    assert result.ac_summary.formulas_restored == 1

    assert result.cumulative_summary.ac_sheet_created == "2월 누계(AC)"
    assert result.cumulative_summary.detail_sheet_created == "2월 누계(상세)"
    assert result.cumulative_summary.prev_sheets_hidden == ["1월 누계(AC)", "1월 누계(상세)"]
    assert result.cumulative_summary.formula_chain_extended == 1

    assert result.verification.verified is True
    assert result.verification.diff == pytest.approx(0.0, abs=1e-6)
    assert result.verification.raw_direct_total == pytest.approx(result.verification.detail_grand_total)
    assert result.verification.derived_cumulative_total == pytest.approx(result.verification.detail_grand_total)

    check_wb = load_workbook(out_path, data_only=False)
    assert check_wb["2월(AC)"].sheet_state == "visible"
    assert check_wb["1월 누계(AC)"].sheet_state == "hidden"
    assert check_wb["1월 누계(상세)"].sheet_state == "hidden"
    assert check_wb["2월 누계(AC)"].sheet_state == "visible"
    assert check_wb["2월 누계(상세)"].sheet_state == "visible"
    assert check_wb["2월(AC)"].cell(2, 2).value == "='2월(BP)'!B2"
    check_wb.close()


async def test_dry_run_does_not_write_or_backup(tmp_path: Path):
    raw_path, template_path = _build_happy_path_fixtures(tmp_path)
    out_path = tmp_path / "out.xlsx"

    result = await update_fc_month_report(
        raw_path=str(raw_path),
        template_path=str(template_path),
        month="2월",
        output_path=str(out_path),
        dry_run=True,
    )

    assert result.dry_run is True
    assert result.output_path is None
    assert result.backup_path is None
    assert not out_path.exists()
    assert result.verification.verified is True


async def test_cross_validate_tolerance_exceeded_raises_verification_failed(tmp_path: Path):
    raw_path, template_path = _build_happy_path_fixtures(tmp_path)
    out_path = tmp_path / "out.xlsx"

    with pytest.raises(errors.VerificationFailed):
        await update_fc_month_report(
            raw_path=str(raw_path),
            template_path=str(template_path),
            month="2월",
            output_path=str(out_path),
            cross_validate_tolerance=-1.0,  # 항상 초과하도록 강제
        )
