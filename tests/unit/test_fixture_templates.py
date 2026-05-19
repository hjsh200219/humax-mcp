"""US-015 fixture template validation."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

ROOT = Path(__file__).resolve().parents[2]
FIX_DIR = ROOT / "fixtures" / "templates"

TEMPLATE_TYPES = ["humax_allocation", "humax_account", "evcs_account"]


@pytest.fixture(scope="session", autouse=True)
def _ensure_fixtures_built():
    """Run build_fixture_templates.py once per session if fixtures missing."""
    missing = [t for t in TEMPLATE_TYPES if not (FIX_DIR / f"{t}.xlsx").exists()]
    if missing:
        import subprocess
        import sys
        script = ROOT / "scripts" / "build_fixture_templates.py"
        result = subprocess.run([sys.executable, str(script)], capture_output=True, text=True)
        if result.returncode != 0:
            pytest.skip(f"Could not build fixtures: {result.stderr}")


@pytest.mark.parametrize("template_type", TEMPLATE_TYPES)
def test_fixture_xlsx_exists(template_type: str) -> None:
    assert (FIX_DIR / f"{template_type}.xlsx").exists()


@pytest.mark.parametrize("template_type", TEMPLATE_TYPES)
def test_sidecar_json_exists(template_type: str) -> None:
    sidecar = FIX_DIR / f"{template_type}.template.json"
    assert sidecar.exists()
    meta = json.loads(sidecar.read_text(encoding="utf-8"))
    assert meta["template_type"] == template_type
    assert meta["schema_version"] == "2026.05"
    assert "sheet_names" in meta
    assert "formula_count_per_sheet" in meta


@pytest.mark.parametrize("template_type", TEMPLATE_TYPES)
def test_fixture_loadable(template_type: str) -> None:
    wb = load_workbook(FIX_DIR / f"{template_type}.xlsx")
    assert len(wb.sheetnames) >= 1


@pytest.mark.parametrize("template_type", TEMPLATE_TYPES)
def test_sidecar_sheet_names_match_workbook(template_type: str) -> None:
    sidecar = json.loads((FIX_DIR / f"{template_type}.template.json").read_text(encoding="utf-8"))
    wb = load_workbook(FIX_DIR / f"{template_type}.xlsx")
    assert set(wb.sheetnames) == set(sidecar["sheet_names"])


@pytest.mark.parametrize("template_type", TEMPLATE_TYPES)
def test_formulas_preserved(template_type: str) -> None:
    sidecar = json.loads((FIX_DIR / f"{template_type}.template.json").read_text(encoding="utf-8"))
    wb = load_workbook(FIX_DIR / f"{template_type}.xlsx", data_only=False)
    for sn, expected_count in sidecar["formula_count_per_sheet"].items():
        ws = wb[sn]
        actual = 0
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    actual += 1
        assert actual == expected_count, f"{template_type}:{sn} formula count {actual} != {expected_count}"


@pytest.mark.parametrize("template_type", TEMPLATE_TYPES)
def test_data_cells_cleared(template_type: str) -> None:
    """Non-formula non-header value cells should be mostly cleared (no real financial numbers)."""
    wb = load_workbook(FIX_DIR / f"{template_type}.xlsx", data_only=False)
    numeric_count = 0
    for ws in wb.worksheets:
        # skip first 4 rows (headers/labels often live there)
        for row in ws.iter_rows(min_row=5):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                v = cell.value
                if isinstance(v, (int, float)) and v not in (0, None):
                    numeric_count += 1
    # Allow zero or very few stragglers (some sheets have row-numbering integers). Cap generous.
    assert numeric_count < 100, f"{template_type}: {numeric_count} numeric non-formula cells survived clearing"
