"""US-008 write_cells tests."""

from __future__ import annotations

import hashlib
from pathlib import Path

import pytest
from openpyxl import load_workbook

from humax_excel_mcp.core import errors
from humax_excel_mcp.schemas.requests import CellUpdate
from humax_excel_mcp.tools.write import write_cells

pytestmark = pytest.mark.asyncio


def _sha(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


async def test_basic_write(sample_26bp_path: Path, tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    res = await write_cells(
        str(sample_26bp_path),
        "예산+실적",
        [CellUpdate(cell="A2", value="EDITED")],
        output_path=str(out),
    )
    assert res.success
    assert res.summary.applied == 1
    assert out.exists()
    wb = load_workbook(out)
    assert wb["예산+실적"]["A2"].value == "EDITED"


async def test_backup_created(sample_26bp_path: Path, tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    res = await write_cells(
        str(sample_26bp_path),
        "예산+실적",
        [CellUpdate(cell="A2", value="X")],
        output_path=str(out),
    )
    assert res.backup_path
    assert Path(res.backup_path).exists()


async def test_overwrite_original_forbidden(sample_26bp_path: Path) -> None:
    with pytest.raises(errors.OverwriteOriginalForbidden):
        await write_cells(
            str(sample_26bp_path),
            "예산+실적",
            [CellUpdate(cell="A2", value="X")],
            output_path=str(sample_26bp_path),
        )


async def test_default_output_path_is_edited_suffix(
    sample_26bp_path: Path,
) -> None:
    res = await write_cells(
        str(sample_26bp_path),
        "예산+실적",
        [CellUpdate(cell="A2", value="X")],
    )
    assert res.output_path is not None
    assert res.output_path.endswith("_edited.xlsx")
    assert Path(res.output_path).exists()


async def test_dry_run_no_file_write(sample_26bp_path: Path, tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    pre_hash = _sha(sample_26bp_path)
    res = await write_cells(
        str(sample_26bp_path),
        "예산+실적",
        [CellUpdate(cell="A2", value="X")],
        output_path=str(out),
        dry_run=True,
    )
    assert res.dry_run is True
    assert res.output_path is None
    assert res.backup_path is None
    assert not out.exists()
    assert _sha(sample_26bp_path) == pre_hash


async def test_skip_formula_default(sample_26bp_path: Path, tmp_path: Path) -> None:
    # Inject a formula manually
    from openpyxl import load_workbook as lw
    wb = lw(sample_26bp_path)
    wb["예산+실적"]["A3"].value = "=SUM(B3:C3)"
    wb.save(sample_26bp_path)

    out = tmp_path / "out.xlsx"
    res = await write_cells(
        str(sample_26bp_path),
        "예산+실적",
        [CellUpdate(cell="A3", value="OVERWRITE", skip_if_formula=True)],
        output_path=str(out),
    )
    assert res.summary.skipped_formula == 1
    assert res.summary.applied == 0


async def test_overwrite_formula_when_disabled(
    sample_26bp_path: Path, tmp_path: Path
) -> None:
    from openpyxl import load_workbook as lw
    wb = lw(sample_26bp_path)
    wb["예산+실적"]["A3"].value = "=SUM(B3:C3)"
    wb.save(sample_26bp_path)
    out = tmp_path / "out.xlsx"
    res = await write_cells(
        str(sample_26bp_path),
        "예산+실적",
        [CellUpdate(cell="A3", value="OVERWRITE", skip_if_formula=False)],
        output_path=str(out),
    )
    assert res.summary.applied == 1
    assert any("수식 덮어쓰기" in w.message for w in res.warnings)


async def test_too_many_updates(sample_26bp_path: Path, tmp_path: Path) -> None:
    updates = [CellUpdate(cell=f"A{i}", value=i) for i in range(1, 5002)]
    with pytest.raises(errors.TooManyUpdates):
        await write_cells(
            str(sample_26bp_path),
            "예산+실적",
            updates,
            output_path=str(tmp_path / "out.xlsx"),
        )


async def test_invalid_cell_address_pydantic() -> None:
    from pydantic import ValidationError
    with pytest.raises(ValidationError):
        CellUpdate(cell="lowercase", value=1)


async def test_file_not_found(tmp_path: Path) -> None:
    with pytest.raises(errors.FileNotFound):
        await write_cells(
            str(tmp_path / "missing.xlsx"),
            "예산+실적",
            [CellUpdate(cell="A2", value=1)],
            output_path=str(tmp_path / "out.xlsx"),
        )


async def test_sheet_not_found(sample_26bp_path: Path, tmp_path: Path) -> None:
    with pytest.raises(errors.SheetNotFound):
        await write_cells(
            str(sample_26bp_path),
            "없는시트",
            [CellUpdate(cell="A2", value=1)],
            output_path=str(tmp_path / "out.xlsx"),
        )


async def test_post_write_verification(sample_26bp_path: Path, tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    res = await write_cells(
        str(sample_26bp_path),
        "예산+실적",
        [CellUpdate(cell="A2", value=42)],
        output_path=str(out),
    )
    assert res.verification.verified is True
    assert res.verification.mismatches == []


async def test_live_artifact_hints(sample_26bp_path: Path, tmp_path: Path) -> None:
    res = await write_cells(
        str(sample_26bp_path),
        "예산+실적",
        [CellUpdate(cell="A2", value="X")],
        output_path=str(tmp_path / "out.xlsx"),
        render_format="live_artifact",
    )
    assert res.artifact_hints is not None
    assert res.artifact_hints.type == "diff_cards"
