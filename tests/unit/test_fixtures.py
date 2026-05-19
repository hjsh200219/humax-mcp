"""US-003 fixture acceptance."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from humax_excel_mcp.schemas import bp26


def test_sample_26bp_path_exists(sample_26bp_path: Path) -> None:
    assert sample_26bp_path.exists()
    assert sample_26bp_path.suffix == ".xlsx"


def test_sample_26bp_has_expected_sheet(sample_26bp_path: Path) -> None:
    wb = load_workbook(sample_26bp_path, data_only=True)
    assert "예산+실적" in wb.sheetnames


def test_sample_26bp_has_full_headers(sample_26bp_path: Path) -> None:
    wb = load_workbook(sample_26bp_path, data_only=True)
    ws = wb["예산+실적"]
    headers = [c.value for c in ws[1]]
    assert "구분" in headers
    assert "Company" in headers
    assert "1월 예산" in headers
    assert "12월 실적" in headers
    assert "STB 배부율" in headers
    assert "Mobility 배부율" in headers
    assert "EVCS국내 배부율" in headers
    assert "EVCS해외 배부율" in headers


def test_sample_26bp_has_data_rows_with_companies(sample_26bp_path: Path) -> None:
    wb = load_workbook(sample_26bp_path, data_only=True)
    ws = wb["예산+실적"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) > 10  # header + many detail rows
    companies = {r[1] for r in rows[1:] if r[1]}
    assert companies & set(bp26.VALID_COMPANIES)


def test_sample_26bp_subtotal_row_present(sample_26bp_path: Path) -> None:
    wb = load_workbook(sample_26bp_path, data_only=True)
    ws = wb["예산+실적"]
    divs = [r[0] for r in ws.iter_rows(values_only=True)]
    assert "총합계" in divs


def test_sample_26bp_allocation_rates_sum_100_on_data_rows(sample_26bp_path: Path) -> None:
    wb = load_workbook(sample_26bp_path, data_only=True)
    ws = wb["예산+실적"]
    headers = [c.value for c in ws[1]]
    rate_idxs = [headers.index(k) for k in bp26.ALLOCATION_RATE_COLUMNS.keys()]
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] == "소조직"]
    assert data_rows
    for r in data_rows[:20]:
        total = sum(float(r[i] or 0) for i in rate_idxs)
        assert abs(total - 100.0) < 0.01, f"rate sum != 100: {total}"
