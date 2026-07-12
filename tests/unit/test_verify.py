"""US-007 verify_sums tests."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from humax_excel_mcp.core import errors
from humax_excel_mcp.tools.verify import verify_sums

pytestmark = pytest.mark.asyncio


async def test_basic_verify(sample_26bp_path: Path) -> None:
    res = await verify_sums(str(sample_26bp_path), "예산+실적")
    assert res.success
    assert res.summary.total_checks >= 1


async def test_total_matches_detail(sample_26bp_path: Path) -> None:
    res = await verify_sums(str(sample_26bp_path), "예산+실적")
    total_check = next((r for r in res.level_results if r.level == "총합계"), None)
    assert total_check is not None
    assert total_check.status == "PASS"


async def test_tolerance_breach_detected(sample_26bp_path: Path, tmp_path: Path) -> None:
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(sample_26bp_path.read_bytes())
    wb = load_workbook(bad)
    ws = wb["예산+실적"]
    headers = [c.value for c in ws[1]]
    annual_actual_idx = headers.index("연간 실적") + 1
    total_row_idx = None
    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == "총합계":
            total_row_idx = ridx
            break
    assert total_row_idx is not None
    ws.cell(row=total_row_idx, column=annual_actual_idx).value = 0
    wb.save(bad)
    res = await verify_sums(str(bad), "예산+실적", tolerance=0.01)
    total_check = next((r for r in res.level_results if r.level == "총합계"), None)
    assert total_check is not None
    assert total_check.status == "FAIL"
    assert res.summary.failed >= 1


async def test_anomaly_detection(sample_26bp_path: Path, tmp_path: Path) -> None:
    bad = tmp_path / "anomaly.xlsx"
    bad.write_bytes(sample_26bp_path.read_bytes())
    wb = load_workbook(bad)
    ws = wb["예산+실적"]
    headers = [c.value for c in ws[1]]
    ab_idx = headers.index("연간 예산") + 1
    aa_idx = headers.index("연간 실적") + 1
    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == "소조직":
            ws.cell(row=ridx, column=ab_idx).value = 1_000_000_000
            ws.cell(row=ridx, column=aa_idx).value = 100_000_000
            break
    wb.save(bad)
    res = await verify_sums(str(bad), "예산+실적")
    assert res.anomalies
    a = res.anomalies[0]
    assert a.flag == "LARGE_VARIANCE"
    assert "백만" in a.suggested_comment


async def test_file_not_found(tmp_path: Path) -> None:
    with pytest.raises(errors.FileNotFound):
        await verify_sums(str(tmp_path / "missing.xlsx"), "예산+실적")


async def test_sheet_not_found(sample_26bp_path: Path) -> None:
    with pytest.raises(errors.SheetNotFound):
        await verify_sums(str(sample_26bp_path), "없는시트")


async def test_live_artifact_hints(sample_26bp_path: Path) -> None:
    res = await verify_sums(str(sample_26bp_path), "예산+실적", render_format="live_artifact")
    assert res.artifact_hints is not None
    assert res.artifact_hints.type == "verification_result"
    assert res.artifact_hints.preferred_chart == "tree"


# --- US-A1: 실계층 rollup 검증 (accuracy-speed-improvement PRD) ---


def _build_hierarchy_workbook(
    path: Path,
    *,
    tamper_l1_subtotal: float = 0.0,
    tamper_detail_annual: float = 0.0,
) -> Path:
    """소조직 상세 3행 + 중조직/대조직 소계 + 총합계를 가진 결정론 fixture."""
    from openpyxl import Workbook

    from tests.conftest import _full_header_row

    headers = _full_header_row()
    wb = Workbook()
    ws = wb.active
    ws.title = "예산+실적"
    ws.append(headers)

    def _append(division, org1, org2, org3, m01_actual, annual_actual):
        vals: dict = {}
        for h in headers:
            vals[h] = 0.0 if ("예산" in h or "실적" in h or "배부율" in h) else ""
        vals.update(
            {
                "구분": division,
                "Company": "HMX",
                "대조직": org1,
                "중조직": org2,
                "소조직": org3,
                "Cost Center": "100001",
                "G/L Account": "511000",
                "G/L Account Name": "급여",
                "1월 실적": m01_actual,
                "연간 실적": annual_actual,
            }
        )
        ws.append([vals[h] for h in headers])

    _append("소조직", "개발", "SW", "1팀", 100.0, 100.0 + tamper_detail_annual)
    _append("소조직", "개발", "SW", "2팀", 150.0, 150.0)
    _append("소조직", "영업", "국내", "1팀", 50.0, 50.0)
    _append("중조직", "개발", "SW", "", 0.0, 250.0)
    _append("대조직", "개발", "", "", 0.0, 250.0 + tamper_l1_subtotal)
    _append("대조직", "영업", "", "", 0.0, 50.0)
    _append("총합계", "", "", "", 0.0, 300.0 + tamper_detail_annual)
    wb.save(path)
    return path


async def test_absent_levels_reported_skipped_not_pass(sample_26bp_path: Path) -> None:
    """소계 행이 없는 계층은 PASS 위장 대신 SKIPPED로 보고해야 한다."""
    res = await verify_sums(str(sample_26bp_path), "예산+실적")
    by_level = {r.level: r for r in res.level_results}
    for lvl in ("사업부", "대조직", "중조직"):
        assert lvl in by_level, f"{lvl} 계층 결과 누락"
        assert by_level[lvl].status == "SKIPPED"


async def test_rollup_all_levels_pass(tmp_path: Path) -> None:
    p = _build_hierarchy_workbook(tmp_path / "h_ok.xlsx")
    res = await verify_sums(str(p), "예산+실적")
    by_level = {r.level: r.status for r in res.level_results}
    assert by_level["총합계"] == "PASS"
    assert by_level["대조직"] == "PASS"
    assert by_level["중조직"] == "PASS"
    assert by_level["소조직"] == "PASS"
    assert by_level["사업부"] == "SKIPPED"


async def test_rollup_detects_wrong_intermediate_subtotal(tmp_path: Path) -> None:
    """대조직 소계를 1원 틀어놓으면 해당 계층 FAIL — 가짜 PASS 금지."""
    p = _build_hierarchy_workbook(tmp_path / "h_bad.xlsx", tamper_l1_subtotal=1.0)
    res = await verify_sums(str(p), "예산+실적", tolerance=0.01)
    l1 = next(r for r in res.level_results if r.level == "대조직")
    assert l1.status == "FAIL"
    assert res.summary.failed >= 1


async def test_detail_annual_vs_monthly_mismatch_fails(tmp_path: Path) -> None:
    """소조직 행의 연간 실적 ≠ 월별 실적 합이면 소조직 계층 FAIL."""
    p = _build_hierarchy_workbook(tmp_path / "h_detail_bad.xlsx", tamper_detail_annual=7.0)
    res = await verify_sums(str(p), "예산+실적", tolerance=0.01)
    detail = next(r for r in res.level_results if r.level == "소조직")
    assert detail.status == "FAIL"


# --- US-A3: 수식 검사 침묵 실패 제거 ---


async def test_formula_check_error_surfaced(sample_26bp_path: Path, monkeypatch) -> None:
    from humax_excel_mcp.core import excel_io as eio

    orig = eio.load_workbook_safe

    def boom(path, *, data_only=True, read_only=False):
        if not data_only:
            raise RuntimeError("formula load boom")
        return orig(path, data_only=data_only, read_only=read_only)

    monkeypatch.setattr(eio, "load_workbook_safe", boom)
    res = await verify_sums(str(sample_26bp_path), "예산+실적", check_formulas=True)
    assert res.metadata.get("formula_check_error")
    assert res.summary.warnings >= 1
