"""US-004 excel_io tests."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from humax_excel_mcp.core import errors, excel_io


def test_assert_xlsx_path_missing(tmp_path: Path) -> None:
    with pytest.raises(errors.FileNotFound):
        excel_io.assert_xlsx_path(tmp_path / "missing.xlsx")


def test_assert_xlsx_path_wrong_ext(tmp_path: Path) -> None:
    p = tmp_path / "data.csv"
    p.write_text("a,b\n1,2", encoding="utf-8")
    with pytest.raises(errors.FileNotFound):
        excel_io.assert_xlsx_path(p)


def test_load_workbook_safe_ok(sample_26bp_path: Path) -> None:
    wb = excel_io.load_workbook_safe(sample_26bp_path)
    assert "예산+실적" in wb.sheetnames


def test_get_sheet_missing(sample_26bp_path: Path) -> None:
    wb = excel_io.load_workbook_safe(sample_26bp_path)
    with pytest.raises(errors.SheetNotFound):
        excel_io.get_sheet(wb, "없는시트")


def test_worksheet_to_dataframe(sample_26bp_path: Path) -> None:
    wb = excel_io.load_workbook_safe(sample_26bp_path)
    ws = excel_io.get_sheet(wb, "예산+실적")
    df = excel_io.worksheet_to_dataframe(ws)
    assert "division" in df.columns
    assert "company" in df.columns
    assert "m01_budget" in df.columns
    assert len(df) > 5


def test_validate_schema_strict_missing_raises(tmp_path: Path) -> None:
    p = tmp_path / "bad.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["구분", "Company"])  # missing tons of cols
    wb.save(p)
    wb2 = excel_io.load_workbook_safe(p)
    ws2 = excel_io.get_sheet(wb2, wb2.sheetnames[0])
    headers = [c.value for c in ws2[1]]
    with pytest.raises(errors.SchemaMismatch):
        excel_io.validate_schema(headers, strict=True)


def test_validate_schema_non_strict_returns_diffs(tmp_path: Path) -> None:
    diffs = excel_io.validate_schema(["구분"], strict=False)
    assert any(d.startswith("MISSING:") for d in diffs)


def test_validate_schema_allocation_cols_required(tmp_path: Path) -> None:
    with pytest.raises(errors.SchemaMismatch):
        excel_io.validate_schema(["구분", "Company"], require_allocation_cols=True)


# US-021 tests — Flexible Header-Row Detection


def test_worksheet_to_dataframe_default_backward_compat(sample_26bp_path: Path) -> None:
    """AC1: Default args use row 1, no behavior change."""
    wb = excel_io.load_workbook_safe(sample_26bp_path)
    ws = excel_io.get_sheet(wb, "예산+실적")
    df = excel_io.worksheet_to_dataframe(ws)
    assert "division" in df.columns
    assert "company" in df.columns
    assert len(df) > 5


def test_worksheet_to_dataframe_explicit_header_row(tmp_path: Path) -> None:
    """AC2: Explicit header_row=3 reads row 3 as header."""
    from openpyxl import Workbook as WB
    p = tmp_path / "real_like.xlsx"
    wb = WB()
    ws = wb.active
    ws.append(["title"])  # row 1
    ws.append(["subtitle"])  # row 2
    ws.append(["Year", "Month", "Company", "Cost Center", "G/L Account", "Amount\n(KRW)", "구분"])  # row 3
    ws.append([2026, 1, "HKR", 101, 500000, 1000, "예산"])  # row 4 data
    ws.append([2026, 2, "HKR", 101, 500000, 2000, "예산"])
    wb.save(p)
    wb2 = excel_io.load_workbook_safe(p)
    ws2 = excel_io.get_sheet(wb2, wb2.sheetnames[0])
    df = excel_io.worksheet_to_dataframe(ws2, header_row=3, schema_module="raw_bp26")
    assert "year" in df.columns
    assert "month" in df.columns
    assert "amount_krw" in df.columns
    assert len(df) == 2


def test_worksheet_to_dataframe_auto_detect_raw(tmp_path: Path) -> None:
    """AC3: auto-detect with schema_module='raw_bp26' finds row 3."""
    from openpyxl import Workbook as WB
    p = tmp_path / "auto.xlsx"
    wb = WB()
    ws = wb.active
    ws.append(["title only"])
    ws.append([None, "sub"])
    ws.append(["Year", "Month", "Company", "Cost Center", "G/L Account", "Amount\n(KRW)", "구분"])
    ws.append([2026, 1, "HKR", 101, 500000, 1000, "예산"])
    wb.save(p)
    wb2 = excel_io.load_workbook_safe(p)
    ws2 = excel_io.get_sheet(wb2, wb2.sheetnames[0])
    df = excel_io.worksheet_to_dataframe(ws2, schema_module="raw_bp26")
    assert "year" in df.columns
    assert len(df) == 1


def test_worksheet_to_dataframe_empty(tmp_path: Path) -> None:
    """AC4: Empty worksheet returns empty DataFrame."""
    from openpyxl import Workbook as WB
    p = tmp_path / "empty.xlsx"
    wb = WB()
    wb.save(p)
    wb2 = excel_io.load_workbook_safe(p)
    ws2 = excel_io.get_sheet(wb2, wb2.sheetnames[0])
    df = excel_io.worksheet_to_dataframe(ws2)
    assert df.empty or len(df) == 0


def test_worksheet_to_dataframe_auto_detect_unknown_raises(tmp_path: Path) -> None:
    """AC5: Auto-detect raises SchemaMismatch when no row matches."""
    from openpyxl import Workbook as WB
    p = tmp_path / "junk.xlsx"
    wb = WB()
    ws = wb.active
    ws.append(["x", "y", "z"])
    ws.append([1, 2, 3])
    wb.save(p)
    wb2 = excel_io.load_workbook_safe(p)
    ws2 = excel_io.get_sheet(wb2, wb2.sheetnames[0])
    with pytest.raises(errors.SchemaMismatch):
        excel_io.worksheet_to_dataframe(ws2, schema_module="raw_bp26")


def test_detect_source_format_aggregated(sample_26bp_path: Path) -> None:
    """detect_source_format -> ('aggregated', 1) for aggregated fixture."""
    wb = excel_io.load_workbook_safe(sample_26bp_path)
    ws = excel_io.get_sheet(wb, "예산+실적")
    fmt, row = excel_io.detect_source_format(ws)
    assert fmt == "aggregated"
    assert row == 1


def test_detect_source_format_raw(tmp_path: Path) -> None:
    """detect_source_format -> ('raw', 3) for row-3-header file."""
    from openpyxl import Workbook as WB
    p = tmp_path / "raw.xlsx"
    wb = WB()
    ws = wb.active
    ws.append(["title"])
    ws.append(["sub"])
    ws.append(["Year", "Month", "Company", "Cost Center", "G/L Account", "Amount\n(KRW)", "구분"])
    ws.append([2026, 1, "HKR", 101, 500000, 1000, "예산"])
    wb.save(p)
    wb2 = excel_io.load_workbook_safe(p)
    ws2 = excel_io.get_sheet(wb2, wb2.sheetnames[0])
    fmt, row = excel_io.detect_source_format(ws2)
    assert fmt == "raw"
    assert row == 3
