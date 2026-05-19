"""Build data-cleared fixture templates from docs/references/*.xlsx.

Reads real Humax output templates, clears all data values while preserving:
- Formulas (cells with value starting with '=')
- Style attributes (font, fill, border, alignment, number_format)
- Merged cells, column/row dimensions
- Sheet structure

Produces:
- fixtures/templates/{humax_allocation,humax_account,evcs_account}.xlsx
- fixtures/templates/{template_type}.template.json (sidecar metadata)

Post-build assertion: per-sheet formula count in fixture == reference.

Usage:
    .venv/bin/python scripts/build_fixture_templates.py
"""

from __future__ import annotations

import json
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

ROOT = Path(__file__).resolve().parents[1]
REF_DIR = ROOT / "docs" / "references"
FIX_DIR = ROOT / "fixtures" / "templates"

KST = timezone(timedelta(hours=9))
SCHEMA_VERSION = "2026.05"

SOURCES: dict[str, str] = {
    "humax_allocation": "Humax FC 실적 (26.03)_260430.xlsx",
    "humax_account": "Humax FC 계정별 실적 (26.03).xlsx",
    "evcs_account": "EVCS FC 계정별 실적 (26.03).xlsx",
}


def count_formulas(ws) -> int:
    n = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                n += 1
    return n


def clear_data_preserve_formulas(ws) -> tuple[int, int]:
    """Clear non-formula values. Returns (cells_cleared, formulas_preserved)."""
    cleared = 0
    preserved = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str) and v.startswith("="):
                preserved += 1
                continue
            cell.value = None
            cleared += 1
    return cleared, preserved


def build_template(template_type: str, source_name: str) -> dict:
    src = REF_DIR / source_name
    if not src.exists():
        raise FileNotFoundError(f"Reference file missing: {src}")

    out_xlsx = FIX_DIR / f"{template_type}.xlsx"
    out_json = FIX_DIR / f"{template_type}.template.json"
    FIX_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(src, keep_vba=False, keep_links=True)

    # Capture reference formula counts before clearing (clearing data should not affect formulas, but capture for safety)
    reference_counts: dict[str, int] = {}
    for sn in wb.sheetnames:
        reference_counts[sn] = count_formulas(wb[sn])

    # Clear data values
    total_cleared = 0
    total_preserved = 0
    for sn in wb.sheetnames:
        c, p = clear_data_preserve_formulas(wb[sn])
        total_cleared += c
        total_preserved += p

    wb.save(out_xlsx)
    wb.close()

    # Post-build assertion: re-open and verify formula counts match reference
    wb_check = load_workbook(out_xlsx, keep_vba=False, keep_links=True)
    fixture_counts: dict[str, int] = {}
    for sn in wb_check.sheetnames:
        fixture_counts[sn] = count_formulas(wb_check[sn])
    wb_check.close()

    for sn in reference_counts:
        if reference_counts[sn] != fixture_counts.get(sn, -1):
            raise ValueError(
                f"{template_type}: formula count mismatch on '{sn}' — "
                f"reference={reference_counts[sn]} fixture={fixture_counts.get(sn)}"
            )

    sidecar = {
        "schema_version": SCHEMA_VERSION,
        "template_type": template_type,
        "source_reference": source_name,
        "sheet_names": list(wb.sheetnames),
        "formula_count_per_sheet": fixture_counts,
        "created_at": datetime.now(KST).isoformat(),
        "stats": {
            "data_cells_cleared": total_cleared,
            "formula_cells_preserved": total_preserved,
        },
    }
    out_json.write_text(
        json.dumps(sidecar, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    return {
        "template_type": template_type,
        "xlsx": str(out_xlsx),
        "sidecar": str(out_json),
        "sheets": len(wb.sheetnames),
        "cleared": total_cleared,
        "preserved": total_preserved,
    }


def main() -> int:
    results: list[dict] = []
    for ttype, source in SOURCES.items():
        try:
            results.append(build_template(ttype, source))
        except (FileNotFoundError, ValueError) as exc:
            print(f"FAIL {ttype}: {exc}", file=sys.stderr)
            return 1
    for r in results:
        print(
            f"OK  {r['template_type']}: {r['sheets']} sheets, "
            f"{r['cleared']} cells cleared, {r['preserved']} formulas preserved"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
